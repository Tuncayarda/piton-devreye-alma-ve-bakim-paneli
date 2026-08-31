#!/usr/bin/env python3
"""Producing the checklist workbook.

The template (Field_Device_Verification.xlsx) is copied, its SOFTWARE CHECK
columns filled from the CURRENT SCAN SNAPSHOT, and saved as a new file. The
template itself is never written to.

Values come from what the panel already shows, not from a fresh network
round: whatever the user saw is what lands in Excel. An unread device leaves
its cell empty — no invented defaults.

Cells greyed out (N/A) in the template mean "not valid for this device type"
and are never written to.

Columns are addressed by their stable id (see `columns`); the heading in the
sheet is resolved to an id once, right after the header row is read.
"""
from __future__ import annotations

from pathlib import Path

from .. import editions, script_loader, settings, status
from ..inventory.device_map import Inventory
from . import columns as cols
from .. import i18n

SHEET_NAME = "Checklist"


def resolve_columns(workbook, verify, sheet: str = SHEET_NAME):
    """The named sheet and its id -> column-index map, validated up front.

    The export and the on-screen preview read the same template, and they
    must fail the same way on the same defect: a missing sheet or a reworded
    required heading is a template problem the operator can fix, so both get
    the translated sentence here instead of one path crashing into a generic
    500. Anything the template carries that this code does not know about
    simply has no id and is left alone.

    Returns ``(worksheet, indexes)``; indexes are 1-based openpyxl columns.
    """
    if sheet not in workbook.sheetnames:
        raise ValueError(i18n.t("error.excelSheetMissing", sheet=sheet))
    worksheet = workbook[sheet]

    indexes = {}
    for index in range(1, worksheet.max_column + 1):
        column = cols.column_for_heading(
            worksheet.cell(verify.HEADER_ROW, index).value)
        if column:
            indexes[column] = index

    required = set(verify.FILLABLE) | {cols.IP_TEMPLATE}
    missing = sorted(column for column in required if column not in indexes)
    if missing:
        raise ValueError(i18n.t("error.excelColumnMissing", columns=missing))
    return worksheet, indexes


def export(inventory: Inventory, results: dict, output: Path | None = None,
           template: Path | None = None,
           sheet: str = SHEET_NAME) -> Path:
    """Produce the workbook and return the path written."""
    import openpyxl

    verify = script_loader.device_verify()   # FILLABLE / HEADER_ROW / NA_FILL
    # The OPEN PROJECT's workbook, not one global file: a stand's
    # devices are not in the trains' template and would fill no rows.
    template = Path(template or editions.checklist_path())
    if not template.exists():
        raise FileNotFoundError(
            i18n.t("error.excelTemplateMissing", path=template))
    # The output goes to the OS Documents folder rather than next to the
    # template, whose directory may be a read-only install location.
    output = Path(output or (
        settings.OUTPUT_DIR
        / f"{template.stem}_set{inventory.set_no}{template.suffix}"))
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.load_workbook(template)
    worksheet, indexes = resolve_columns(workbook, verify, sheet)

    by_template = {device.ip_template: device for device in inventory.devices}
    for row in range(verify.HEADER_ROW + 1, worksheet.max_row + 1):
        template_ip = worksheet.cell(row, indexes[cols.IP_TEMPLATE]).value
        device = by_template.get(str(template_ip)) if template_ip else None
        if device is None:
            continue

        values = {cols.DEVICE_NAME: device.name}
        result = results.get(device.id)
        if result is not None and result.state == status.OK:
            values[cols.CONNECTION_INFO] = device.ip
            values[cols.STATUS_DESCRIPTION] = cols.STATUS_ACTIVE
            for field, column in cols.COLUMN_FOR_FIELD.items():
                value = result.fields.get(field)
                if value not in (None, ""):
                    values[column] = value
        elif result is not None and result.state in (status.FAILED,
                                                     status.AUTH):
            values[cols.STATUS_DESCRIPTION] = cols.STATUS_INACTIVE

        for column, value in values.items():
            if column not in verify.FILLABLE or value in (None, ""):
                continue
            cell = worksheet.cell(row, indexes[column])
            if (cell.fill and cell.fill.fgColor
                    and cell.fill.fgColor.rgb == verify.NA_FILL):
                continue                      # grey = invalid for this type
            cell.value = value

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output
