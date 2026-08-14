#!/usr/bin/env python3
"""On-screen preview of the checklist workbook.

The table shown on screen is an exact preview of the Excel that will be
produced: column headings, group headings (PHYSICAL CHECK / VERIFICATION
CRITERIA / SOFTWARE CHECK), section separators and every greyed-out (N/A)
cell are read from the template file. No separate column list is kept — change
the template and the screen follows, so the two cannot drift apart.

The SOFTWARE CHECK columns are filled from the current scan snapshot;
workbook.py uses the same mapping, so what is on screen is what is written.

Each column carries its stable id alongside the heading (see `columns`), so
the code below never has to compare a heading string.
"""
from __future__ import annotations

import threading
from pathlib import Path

from .. import script_loader, settings, status
from ..inventory.device_map import Inventory
from . import columns as cols
from .workbook import SHEET_NAME
from .. import i18n

# The template is not re-parsed on every request; it refreshes when the file
# changes.
_CACHE: dict[tuple, dict] = {}
_LOCK = threading.Lock()


def _merged_width(worksheet, row: int, column: int) -> int:
    """How many columns a cell spans (1 when not merged)."""
    for merged in worksheet.merged_cells.ranges:
        if (merged.min_row <= row <= merged.max_row
                and merged.min_col <= column <= merged.max_col):
            return merged.max_col - merged.min_col + 1
    return 1


def _spans_full_row(worksheet, row: int, column_count: int) -> bool:
    return _merged_width(worksheet, row, 1) >= column_count


def template_layout(template: Path | None = None,
                    sheet: str = SHEET_NAME) -> dict:
    """The template's skeleton: groups, columns, sections and the N/A map."""
    import openpyxl

    template = Path(template or settings.EXCEL_TEMPLATE)
    if not template.exists():
        raise FileNotFoundError(
            i18n.t("error.excelTemplateMissing", path=template))
    stamp = (str(template), template.stat().st_mtime_ns)
    with _LOCK:
        ready = _CACHE.get(stamp)
        if ready:
            return ready

    verify = script_loader.device_verify()
    workbook = openpyxl.load_workbook(template)
    worksheet = workbook[sheet]
    column_count = worksheet.max_column
    header_row = verify.HEADER_ROW

    # ── group headings (one row above the header row) ──
    groups = []
    column = 1
    while column <= column_count:
        width = _merged_width(worksheet, header_row - 1, column)
        value = worksheet.cell(header_row - 1, column).value
        groups.append({"name": str(value or ""), "width": width})
        column += width

    # ── column headings ──
    columns = []
    for column in range(1, column_count + 1):
        name = worksheet.cell(header_row, column).value
        column_id = cols.column_for_heading(name)
        columns.append({
            "id": column_id,
            "name": str(name or ""),
            "manual": column_id in cols.MANUAL_COLUMNS,
            "templateKey": column_id == cols.IP_TEMPLATE,
        })

    # ── sections and rows ──
    sections: list[dict] = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        if _spans_full_row(worksheet, row, column_count):
            heading = str(worksheet.cell(row, 1).value or "").strip()
            if heading and heading.upper() != cols.SUMMARY_HEADING:
                sections.append({"heading": heading, "rows": []})
            elif heading.upper() == cols.SUMMARY_HEADING:
                break                      # the formula summary is not previewed
            continue
        if not sections:
            continue
        key_index = next(index for index, column in enumerate(columns)
                         if column["templateKey"])
        template_ip = worksheet.cell(row, key_index + 1).value
        if not template_ip:
            continue
        cells = []
        for column in range(1, column_count + 1):
            cell = worksheet.cell(row, column)
            not_applicable = bool(cell.fill and cell.fill.fgColor
                                  and cell.fill.fgColor.rgb == verify.NA_FILL)
            value = cell.value
            cells.append({
                "templateValue": "" if value is None else str(value),
                "notApplicable": not_applicable,
            })
        sections[-1]["rows"].append({
            "ipTemplate": str(template_ip),
            "cells": cells,
        })

    layout = {"groups": groups, "columns": columns, "sections": sections,
              "file": template.name}
    with _LOCK:
        _CACHE[stamp] = layout
    return layout


def build_preview(inventory: Inventory, results: dict,
                  template: Path | None = None) -> dict:
    """Fill the template skeleton with the current scan snapshot."""
    layout = template_layout(template)
    by_template = {device.ip_template: device for device in inventory.devices}

    sections = []
    for section in layout["sections"]:
        rows = []
        for raw in section["rows"]:
            device = by_template.get(raw["ipTemplate"])
            result = results.get(device.id) if device else None
            values = (result.fields if result else {}) or {}
            read_ok = result is not None and result.state == status.OK

            cells = []
            for index, cell in enumerate(raw["cells"]):
                column = layout["columns"][index]["id"]
                value = cell["templateValue"]
                source = "template"

                if column == cols.EXPECTED_IP and device:
                    value, source = device.ip, "computed"
                elif column == cols.EXPECTED_SIP_EXTENSION and device:
                    value = device.pbx_extension or value
                elif (not cell["notApplicable"]
                        and column and column not in cols.MANUAL_COLUMNS):
                    source = "probe"
                    if column == cols.DEVICE_NAME:
                        value = device.name if device else ""
                    elif column == cols.CONNECTION_INFO:
                        value = device.ip if (read_ok and device) else ""
                    elif column == cols.STATUS_DESCRIPTION:
                        value = (cols.STATUS_ACTIVE if read_ok
                                 else (cols.STATUS_INACTIVE if result else ""))
                    else:
                        field = cols.FIELD_FOR_COLUMN.get(column)
                        raw_value = values.get(field) if field else None
                        value = ("" if raw_value in (None, "")
                                 else str(raw_value))

                cells.append({
                    "value": value,
                    "notApplicable": cell["notApplicable"],
                    "source": source,
                })

            rows.append({
                "deviceId": device.id if device else None,
                "ipTemplate": raw["ipTemplate"],
                "ip": device.ip if device else "",
                "name": device.name if device else "",
                "state": result.state if result else status.UNKNOWN,
                "detail": result.detail if result else i18n.t("probe.notReadYet"),
                "cells": cells,
            })
        if rows:
            sections.append({"heading": section["heading"], "rows": rows})

    return {
        "file": layout["file"],
        "output": f"{Path(layout['file']).stem}_set{inventory.set_no}.xlsx",
        "groups": layout["groups"],
        "columns": layout["columns"],
        "sections": sections,
    }


def clear_cache() -> None:
    with _LOCK:
        _CACHE.clear()
