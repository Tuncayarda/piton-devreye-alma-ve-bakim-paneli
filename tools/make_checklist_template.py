#!/usr/bin/env python3
"""Build a project's checklist workbook from its DeviceMap.

    python3 tools/make_checklist_template.py fuar

WHY A TOOL AND NOT A HAND-MADE FILE. The workbook matches its rows to devices
by IP template (`panel/checklist/workbook.py`), and the greyed-out cells in it
are what the on-screen preview reads to decide which fields a device type even
has (`panel/checklist/preview.py`). A file typed by hand drifts from the
DeviceMap the first time a device moves port, and the drift is silent: the row
simply stops being filled.

THE STYLE COMES FROM THE YATAKLI TEMPLATE, which is the one that was designed.
Its header rows are copied verbatim, and each device row takes its formatting
from a row of the same device type there — including the grey N/A fills, which
carry the real knowledge in this file: which columns a PISCU has and a camera
has not. Nothing about that is re-decided here.

A type the base template has never seen falls back to another with the SAME
READ METHOD, because that is what decides which fields exist. The fallback is
printed, so a new device type shows up in the output rather than quietly
inheriting the wrong columns.
"""
from __future__ import annotations

import shutil
import sys
from copy import copy
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl                                        # noqa: E402

from panel.editions import catalogue                   # noqa: E402
from panel.inventory import catalog, device_map        # noqa: E402

BASE = (ROOT / catalogue.MAPS_DIR / catalogue.BASE_DIR
        / catalogue.BASE_CHECKLIST)
SHEET = "Checklist"
HEADER_ROWS = 4                     # set-no, title, groups, column headings
BANNER_FILL = "FFBDD7EE"


def _key(device) -> tuple[str, str]:
    return (device.type, device.subtype or "")


def _label(type_name: str, subtype: str) -> str:
    return f"{type_name} · {subtype}" if subtype else type_name


def donors(worksheet) -> list[tuple[str, str, int]]:
    """(type, subtype, row) read out of the base template's banners.

    Read from the file rather than listed here: the template is the source of
    truth for which columns a type has — the grey cells ARE that knowledge —
    and a list beside it would be a second one to keep in step.

    Matched case-insensitively later: the banners are shouted ("SWITCH") and
    a DeviceMap is not ("Switch").
    """
    found: list[tuple[str, str, int]] = []
    section = None
    for row in range(HEADER_ROWS + 1, worksheet.max_row + 1):
        cell = worksheet.cell(row, 1)
        fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else ""
        if cell.value and fill == BANNER_FILL:
            head = str(cell.value).split("·")[0:2]
            name = head[0].strip()
            sub = (head[1].strip()
                   if len(head) > 1 and "record" not in head[1] else "")
            section = (name, sub)
            continue
        known = any(t == section[0] and s == section[1] for t, s, _r in found)
        if cell.value and section and not known:
            found.append((section[0], section[1], row))
    return found


# ── the version each device kind is expected to be running ───────────────
#
# Column G. It is a MANUAL column (`panel/checklist/columns.py`): the panel
# reads it out of this workbook and never computes it, so an empty cell means
# the commissioning record ships with nothing to compare the reading against.
# All five workbooks shipped that way until this table was written.
#
# The values are the ones `field_scripts/device_verify.py` has been checking
# against on the trains. They are written HERE rather than in `panel/` because
# nothing at run time uses them: they are baked into a generated file, and a
# firmware release is a re-run of this tool rather than a code change.
#
# There is no verification FORMULA to write. The old field script's workbook
# carried a "version check" column beside this one; the panel's does not —
# it records what was expected and what was read, side by side, and leaves
# the judgement to the person signing the sheet.
VERSIONS_EVERYWHERE = {
    # The two long-range cameras run a different firmware from the rest, so
    # they are named; every other camera falls to the ("Camera", "") row.
    ("Camera", "Rear"): "V4.3.1build260109 736311",
    ("Camera", "Panto"): "V4.3.1build260109 736311",
    ("Camera", ""): "V1.4.1build260128 741107",
    # The Android displays report the PACKAGE and its version, not a build
    # id — all four of them, because they are the same hardware running the
    # same application (see `panel/inventory/catalog.py`). Listed one by one
    # rather than falling back to ("LCD", ""), which is the passive panel's
    # own firmware and would be the wrong thing to expect of an app.
    ("LCD", "Compartment"): "com.piton.train_lcd_panel 0.0.5",
    ("LCD", "Twin"): "com.piton.train_lcd_panel 0.0.5",
    ("LCD", "LINE"): "com.piton.train_lcd_panel 0.0.5",
    ("LCD", "PIS"): "com.piton.train_lcd_panel 0.0.5",
    ("LCD", "Landing"): "com.piton.train_lcd_panel 0.0.5",
    ("LCD", ""): "1.0.0",
    ("NVR", ""): "V5.4.1.630386build250318",
    # Two that are not a single version. The sheet is read by a person, so
    # they are written the way the answer is actually given.
    ("LED", ""): "7 or newer",
    ("Switch", ""): "R2008 or R1002",
    ("Router", ""): "8.2.0 build 4901",
}

# What each project runs on top of that. The announcement equipment is the
# real split: Gaziray and GDM run a version per subtype, while the other
# three run one build across the whole family.
VERSIONS_BY_PROJECT = {
    "yatakli": {("Announcement", ""): "1.2.8",
                ("PISCU", ""): "1.2.7", ("HMI", ""): "1.2.7",
                ("ICU", ""): "1.2.7"},
    "vip": {("Announcement", ""): "1.2.8",
            ("PISCU", ""): "1.2.7", ("HMI", ""): "1.2.7",
            ("ICU", ""): "1.2.7"},
    "gaziray": {("Announcement", "Intercom"): "5.0.9.2",
                ("Announcement", "Swanneck"): "5.0.9",
                ("Announcement", "Amplifier"): "5.0.8",
                ("Announcement", "UIC"): "1.2.7",
                ("PISCU", ""): "1.6.1", ("HMI", ""): "1.6.1"},
    "gdm": {("Announcement", "Intercom"): "5.0.9.2",
            ("Announcement", "Swanneck"): "5.0.9",
            ("Announcement", "Amplifier"): "5.0.8",
            ("Announcement", "UIC"): "1.2.7",
            ("PISCU", ""): "2.0.4", ("HMI", ""): "2.0.4"},
    "fuar": {("Announcement", ""): "1.2.8",
             ("PISCU", ""): "1.0.0", ("HMI", ""): "1.0.0"},
}


def expected_version(project_key: str, device) -> str:
    """The version this device should be running, or "" when none is stated.

    Exact subtype first, then the type's own row. NOT a read-method fallback,
    which is right for the donor row above and wrong here: an LCD·LINE and an
    LCD·Compartment share a read method AND a version, but a Camera·Corridor
    and a Camera·Rear share the method and run different firmware.

    Blank is a real answer. GDM's induction loop has no API and reports
    nothing; a version expected of it would be a cell nobody could fill.
    """
    table = {**VERSIONS_EVERYWHERE,
             **VERSIONS_BY_PROJECT.get(project_key, {})}
    subtype = device.subtype or ""
    if (device.type, subtype) in table:
        return table[(device.type, subtype)]
    return table.get((device.type, ""), "")


# Decided by hand, with the reason, where the rules below cannot tell.
# A gooseneck microphone is a microphone: it reports the same gains and the
# same outbound number a handset does, and the rules would otherwise hand it
# the UIC row, which has neither.
BY_HAND = {("announcement", "swanneck"): ("Announcement", "Handset")}


def pick_donor(device, table, worksheet) -> tuple[int, str]:
    """The row to copy formatting from, and a note when it is not exact.

    THE READ METHOD OUTRANKS THE SUBTYPE, because it is what decides which
    fields a device has at all. An LCD·Twin is read over ADB and an
    LCD·Landing over MQTT; taking the Landing row because both are LCDs would
    grey out the version, timezone and SIP columns the Twin actually fills.
    """
    label = _label(device.type, device.subtype or "")
    want = (device.type.lower(), (device.subtype or "").lower())
    method = catalog.read_method_for(device.type, device.subtype or "")

    def find(match):
        for name, sub, row in table:
            if match(name, sub):
                return name, sub, row
        return None

    exact = find(lambda n, s: (n.lower(), s.lower()) == want)
    if exact:
        return exact[2], ""

    chosen = BY_HAND.get(want)
    if chosen:
        found = find(lambda n, s: (n, s) == chosen)
        if found:
            return found[2], f"{label}: the {_label(*chosen)} row — decided by hand"

    for describe, match in (
            ("same type, same read method",
             lambda n, s: n.lower() == want[0]
             and catalog.read_method_for(n, s) == method),
            (f"same read method ({method})",
             lambda n, s: catalog.read_method_for(n, s) == method),
            ("same type", lambda n, s: n.lower() == want[0])):
        found = find(match)
        if found:
            return found[2], (f"{label}: the {_label(found[0], found[1])} "
                              f"row — {describe}")
    raise SystemExit(f"[tool] no row to copy formatting from: {label}")


def build(project_key: str) -> Path:
    project = next((p for p in catalogue.ALL_PROJECTS
                    if p.key == project_key), None)
    if project is None:
        raise SystemExit(f"[tool] unknown project: {project_key}. Choices: "
                         f"{', '.join(p.key for p in catalogue.ALL_PROJECTS)}")
    source = ROOT.joinpath(*project.source_path)
    if not source.exists():
        raise SystemExit(f"[tool] no DeviceMap at {source}")

    inventory = device_map.load(1, path=source, cache=False)

    # THE BASE FILE IS COPIED AND EDITED, not rebuilt beside it. openpyxl
    # stores a cell's formatting as an index into the workbook's own style
    # tables, so a style carried across two workbooks points at nothing —
    # which is not an error until the file is opened. Editing one workbook
    # keeps every index valid, and the header rows, column widths, frozen
    # pane and print setup come along untouched.
    #
    # Beside the DeviceMap it was built from, in that project's folder, and
    # the table is what says where: the naming rule lives in one place
    # (`panel/editions/catalogue.py`) and this tool applies it rather than
    # spelling a second name out. The two disagreeing is a package that
    # builds and then cannot find its own workbook.
    out_path = ROOT.joinpath(*project.checklist_source)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(BASE, out_path)
    book = openpyxl.load_workbook(out_path)
    sheet = book[SHEET]
    columns = sheet.max_column
    table = donors(sheet)

    # Devices grouped by type, in the order the DeviceMap lists them: the
    # operator walks the rack in that order and the sheet should follow.
    groups: dict[tuple[str, str], list] = {}
    for device in inventory.devices:
        groups.setdefault(_key(device), []).append(device)

    def row_style(row_number: int) -> list:
        """One row's formatting, lifted out of the sheet.

        TAKEN BEFORE A SINGLE CELL IS WRITTEN, and that is the whole point.
        The body is rewritten IN PLACE from the top down, so by the time the
        fifth section is written the rows it wanted to copy from have already
        been overwritten by the first four — a donor at row 38 is not the
        Amplifier any more, it is whatever section landed there.

        Nothing said so. The rows came out with a banner's fill on a device
        and a device's fill on a banner, which reads as a cosmetic slip and
        is not one: the grey N/A cells ARE the record of which fields a type
        has (`panel/checklist/preview.py` reads them), so a shifted row tells
        the panel an Intercom has no speaker volume. The Fuar workbook in the
        field has carried exactly that since it was generated.
        """
        return [copy(sheet.cell(row_number, column)._style)
                for column in range(1, columns + 1)]

    plan, notes = [], []
    for key, devices in groups.items():
        donor, note = pick_donor(devices[0], table, sheet)
        if note:
            notes.append(note)
        # The banner is the row above the section's first data row.
        plan.append((key, devices, row_style(donor), row_style(donor - 1)))

    # Every body merge goes before anything is written: the banners are
    # full-width merges and a leftover one would swallow a device row.
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row > HEADER_ROWS:
            sheet.unmerge_cells(str(merged))

    def paste_row(styles: list, into: int) -> None:
        for column, style in enumerate(styles, start=1):
            cell = sheet.cell(into, column)
            cell._style = copy(style)
            cell.value = None

    row = HEADER_ROWS + 1
    banners = []
    for key, devices, donor, banner in plan:
        paste_row(banner, row)
        sheet.cell(row, 1).value = f"  {_label(*key)}   ·   {len(devices)} records"
        banners.append(row)
        row += 1
        for device in devices:
            paste_row(donor, row)
            sheet.cell(row, 1).value = project.label
            sheet.cell(row, 2).value = device.switch_name
            sheet.cell(row, 3).value = device.port
            sheet.cell(row, 4).value = device.name
            sheet.cell(row, 5).value = device.ip_template
            sheet.cell(row, 6).value = (
                f'=IF($B$1="","",SUBSTITUTE(E{row},"n",$B$1))')
            sheet.cell(row, 7).value = (
                expected_version(project.key, device) or None)
            row += 1

    if sheet.max_row >= row:
        sheet.delete_rows(row, sheet.max_row - row + 1)
    for banner in banners:
        sheet.merge_cells(start_row=banner, start_column=1,
                          end_row=banner, end_column=columns)
    sheet.cell(2, 1).value = (
        f"{project.label.upper()} — FIELD DEVICE VERIFICATION")

    book.save(out_path)
    print(f"[tool] {out_path.relative_to(ROOT)} — {len(inventory.devices)} "
          f"devices, {len(groups)} sections, {row - 1} rows")
    for note in notes:
        print(f"[tool] formatting borrowed — {note}")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit(__doc__)
    build(sys.argv[1])
