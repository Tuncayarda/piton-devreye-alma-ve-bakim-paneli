#!/usr/bin/env python3
"""Kontrol listesi — Excel şablonunun ekran önizlemesi.

Ekranda gösterilen tablo, üretilecek Excel'in birebir önizlemesidir:
sütun başlıkları, grup başlıkları (FİZİKSEL KONTROL / DOĞRULAMA KRİTERİ /
YAZILIM KONTROL), bölüm ayırıcıları ve gri (N/A) hücrelerin tamamı
şablon dosyasından okunur. Ayrı bir kolon listesi tutulmaz — şablon
değişince ekran da kendiliğinden değişir, ikisi ayrışmaz.

"YAZILIM KONTROL" sütunları o anki tarama görüntüsünden doldurulur;
core/excel.py aynı eşlemeyi kullanır, dolayısıyla ekranda görünen ile
dosyaya yazılan aynıdır.
"""
from __future__ import annotations

import threading
from pathlib import Path

from . import ayar, betik, dogrulama
from .device_map import Envanter
from .excel import ESLEME

# Şablon her istekte yeniden ayrıştırılmaz; dosya değişirse tazelenir.
_ONBELLEK: dict[tuple, dict] = {}
_KILIT = threading.Lock()

# Şablonda elle doldurulan, cihazdan okunmayan sütunlar
ELLE = {"Bölüm", "Switch", "Port", "Cihaz Tanımı", "IP Şablonu",
        "Beklenen IP", "Beklenen Versiyon", "Beklenen SIP Dahili No"}


def _birlesik_genislik(ws, satir: int, sutun: int) -> int:
    """Bir hücrenin kaç sütuna yayıldığı (birleşik değilse 1)."""
    for aralik in ws.merged_cells.ranges:
        if (aralik.min_row <= satir <= aralik.max_row
                and aralik.min_col <= sutun <= aralik.max_col):
            return aralik.max_col - aralik.min_col + 1
    return 1


def _tum_satiri_kaplar(ws, satir: int, sutun_sayisi: int) -> bool:
    return _birlesik_genislik(ws, satir, 1) >= sutun_sayisi


def sablon_yapisi(sablon: Path | None = None,
                  sayfa: str = "Kontrol Listesi") -> dict:
    """Şablonun iskeletini çıkarır: gruplar, sütunlar, bölümler, N/A haritası."""
    import openpyxl

    sablon = Path(sablon or ayar.EXCEL_SABLON)
    if not sablon.exists():
        raise FileNotFoundError(f"Excel şablonu bulunamadı: {sablon}")
    imza = (str(sablon), sablon.stat().st_mtime_ns)
    with _KILIT:
        hazir = _ONBELLEK.get(imza)
        if hazir:
            return hazir

    dv = betik.device_verify()
    wb = openpyxl.load_workbook(sablon)
    ws = wb[sayfa]
    n_sutun = ws.max_column
    baslik_satiri = dv.HEADER_ROW

    # ── grup başlıkları (şablonda başlık satırının bir üstü) ──
    gruplar = []
    c = 1
    while c <= n_sutun:
        genislik = _birlesik_genislik(ws, baslik_satiri - 1, c)
        deger = ws.cell(baslik_satiri - 1, c).value
        gruplar.append({"ad": str(deger or ""), "genislik": genislik})
        c += genislik

    # ── sütun başlıkları ──
    sutunlar = []
    for c in range(1, n_sutun + 1):
        ad = ws.cell(baslik_satiri, c).value
        sutunlar.append({
            "ad": str(ad or ""),
            "elle": str(ad or "") in ELLE,
            "sablonAlan": str(ad or "") == dv.COL_IP_TEMPLATE,
        })

    # ── bölümler ve satırlar ──
    bolumler: list[dict] = []
    for r in range(baslik_satiri + 1, ws.max_row + 1):
        if _tum_satiri_kaplar(ws, r, n_sutun):
            baslik = str(ws.cell(r, 1).value or "").strip()
            if baslik and baslik.upper() != "SONUÇ":
                bolumler.append({"baslik": baslik, "satirlar": []})
            elif baslik.upper() == "SONUÇ":
                break                      # formül özeti önizlemede yok
            continue
        if not bolumler:
            continue
        sablon_ip = ws.cell(r, sutunlar.index(
            next(s for s in sutunlar if s["sablonAlan"])) + 1).value
        if not sablon_ip:
            continue
        hucreler = []
        for c in range(1, n_sutun + 1):
            hucre = ws.cell(r, c)
            na = bool(hucre.fill and hucre.fill.fgColor
                      and hucre.fill.fgColor.rgb == dv.NA_FILL)
            deger = hucre.value
            hucreler.append({
                "sablonDeger": "" if deger is None else str(deger),
                "na": na,
            })
        bolumler[-1]["satirlar"].append({
            "ipSablonu": str(sablon_ip),
            "hucreler": hucreler,
        })

    yapi = {"gruplar": gruplar, "sutunlar": sutunlar, "bolumler": bolumler,
            "dosya": sablon.name}
    with _KILIT:
        _ONBELLEK[imza] = yapi
    return yapi


def onizleme(env: Envanter, sonuclar: dict,
             sablon: Path | None = None) -> dict:
    """Şablon iskeletini o anki tarama görüntüsüyle doldurur."""
    yapi = sablon_yapisi(sablon)
    sablon_ile = {c.ip_sablonu: c for c in env.cihazlar}
    # Excel sütun başlığı -> sonuç alanı (ESLEME'nin tersi)
    baslik_alan = {baslik: alan for alan, baslik in ESLEME.items()}

    bolumler = []
    for bolum in yapi["bolumler"]:
        satirlar = []
        for ham in bolum["satirlar"]:
            cihaz = sablon_ile.get(ham["ipSablonu"])
            sonuc = sonuclar.get(cihaz.id) if cihaz else None
            alanlar = (sonuc.alanlar if sonuc else {}) or {}
            okundu = sonuc is not None and sonuc.durum == dogrulama.YESIL

            hucreler = []
            for i, hucre in enumerate(ham["hucreler"]):
                baslik = yapi["sutunlar"][i]["ad"]
                deger = hucre["sablonDeger"]
                kaynak = "sablon"

                if baslik == "Beklenen IP" and cihaz:
                    deger, kaynak = cihaz.ip, "hesaplanan"
                elif baslik == "Beklenen SIP Dahili No" and cihaz:
                    deger = cihaz.pbx_extension or deger
                elif not hucre["na"] and baslik not in ELLE:
                    kaynak = "okuma"
                    if baslik == "Cihaz İsmi":
                        deger = cihaz.ad if cihaz else ""
                    elif baslik == "Bağlantı Bilgisi":
                        deger = cihaz.ip if (okundu and cihaz) else ""
                    elif baslik == "Durum Açıklaması":
                        deger = ("Aktif" if okundu
                                 else ("Pasif" if sonuc else ""))
                    else:
                        alan = baslik_alan.get(baslik)
                        ham_deger = alanlar.get(alan) if alan else None
                        deger = "" if ham_deger in (None, "") else str(ham_deger)

                hucreler.append({
                    "deger": deger,
                    "na": hucre["na"],
                    "kaynak": kaynak,
                })

            satirlar.append({
                "cihazId": cihaz.id if cihaz else None,
                "ipSablonu": ham["ipSablonu"],
                "ip": cihaz.ip if cihaz else "",
                "ad": cihaz.ad if cihaz else "",
                "durum": sonuc.durum if sonuc else dogrulama.GRI,
                "aciklama": sonuc.aciklama if sonuc else "Henüz okunmadı",
                "hucreler": hucreler,
            })
        if satirlar:
            bolumler.append({"baslik": bolum["baslik"], "satirlar": satirlar})

    return {
        "dosya": yapi["dosya"],
        "cikti": f"{Path(yapi['dosya']).stem}_set{env.set_no}.xlsx",
        "gruplar": yapi["gruplar"],
        "sutunlar": yapi["sutunlar"],
        "bolumler": bolumler,
    }


def onbellegi_bosalt() -> None:
    with _KILIT:
        _ONBELLEK.clear()
