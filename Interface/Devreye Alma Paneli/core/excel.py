#!/usr/bin/env python3
"""Kontrol listesi Excel'i.

Şablon (Yatakli_Saha_Cihaz_Dogrulama.xlsx) kopyalanır, "YAZILIM KONTROL"
sütunları o anki TARAMA GÖRÜNTÜSÜNDEN doldurulur ve yeni dosya olarak
kaydedilir. Şablon dosyasına asla yazılmaz.

Değerler yeniden ağa çıkılarak değil, panelin ekranda gösterdiği
sonuçlardan alınır: kullanıcı neyi gördüyse Excel'de de o vardır. Cihaz
okunamadıysa hücre boş bırakılır — sahte varsayılan yazılmaz.

Şablonda gri (N/A) boyanmış hücreler o cihaz türünde geçersiz demektir;
oralara hiçbir koşulda yazılmaz.
"""
from __future__ import annotations

from pathlib import Path

from . import ayar, betik, dogrulama
from .device_map import Envanter

# Sonuç alanı -> Excel sütun başlığı
ESLEME = {
    "surum": "Versiyon",
    "seri": "Cihaz Numarası",
    "calisma": "Çalışma Süresi",
    "hoparlor": "Hoparlör Ses Seviyesi",
    "mikrofon": "Mikrofon Ses Seviyesi",
    "hoparlorGain": "Hoparlör Gain",
    "mikrofonGain": "Mikrofon Gain",
    "sipPbx": "SIP PBX IP",
    "sipDahili": "SIP Dahili No",
    "sipArama": "SIP Arama No",
    "saatDilimi": "Saat Dilimi",
    "agZaman": "Ağ/Zaman Kontrolü",
}


def uret(env: Envanter, sonuclar: dict, cikti: Path | None = None,
         sablon: Path | None = None, sayfa: str = "Kontrol Listesi") -> Path:
    """Excel'i üretir ve yazılan dosyanın yolunu döndürür."""
    import openpyxl

    dv = betik.device_verify()          # FILLABLE / HEADER_ROW / NA_FILL
    sablon = Path(sablon or ayar.EXCEL_SABLON)
    if not sablon.exists():
        raise FileNotFoundError(f"Excel şablonu bulunamadı: {sablon}")
    # Çıktı, şablonun yanına değil işletim sisteminin Belgeler klasörüne
    # yazılır: şablonun bulunduğu dizin kurulum dizini olabiliyor.
    cikti = Path(cikti or (ayar.CIKTI_DIZINI
                           / f"{sablon.stem}_set{env.set_no}{sablon.suffix}"))
    cikti.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(sablon)
    if sayfa not in wb.sheetnames:
        raise ValueError(f"Şablonda '{sayfa}' sayfası yok")
    ws = wb[sayfa]

    sutun = {ws.cell(dv.HEADER_ROW, c).value: c
             for c in range(1, ws.max_column + 1)
             if ws.cell(dv.HEADER_ROW, c).value}
    eksik = [b for b in dv.FILLABLE + [dv.COL_IP_TEMPLATE] if b not in sutun]
    if eksik:
        raise ValueError(f"Şablonda bulunamayan sütun: {eksik}")

    sablon_ile = {c.ip_sablonu: c for c in env.cihazlar}
    yazilan = 0
    for satir in range(dv.HEADER_ROW + 1, ws.max_row + 1):
        tmpl = ws.cell(satir, sutun[dv.COL_IP_TEMPLATE]).value
        cihaz = sablon_ile.get(str(tmpl)) if tmpl else None
        if cihaz is None:
            continue

        degerler = {"Cihaz İsmi": cihaz.ad}
        sonuc = sonuclar.get(cihaz.id)
        if sonuc is not None and sonuc.durum == dogrulama.YESIL:
            degerler["Bağlantı Bilgisi"] = cihaz.ip
            degerler["Durum Açıklaması"] = "Aktif"
            for anahtar, baslik in ESLEME.items():
                deger = sonuc.alanlar.get(anahtar)
                if deger not in (None, ""):
                    degerler[baslik] = deger
        elif sonuc is not None and sonuc.durum in (dogrulama.KIRMIZI,
                                                   dogrulama.TURUNCU):
            degerler["Durum Açıklaması"] = "Pasif"

        for baslik, deger in degerler.items():
            if baslik not in dv.FILLABLE or deger in (None, ""):
                continue
            hucre = ws.cell(satir, sutun[baslik])
            if (hucre.fill and hucre.fill.fgColor
                    and hucre.fill.fgColor.rgb == dv.NA_FILL):
                continue                      # gri = bu tür için geçersiz
            hucre.value = deger
            yazilan += 1

    cikti.parent.mkdir(parents=True, exist_ok=True)
    wb.save(cikti)
    return cikti
