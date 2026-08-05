#!/usr/bin/env python3
"""Kontrol listesi önizlemesi ve canlı işlem göstergesi.

Bu ekran şablonun birebir önizlemesi olmak zorunda: sütunlar, gruplar ve
N/A hücreleri şablondan gelmeli, ekranda görünen değer Excel'e yazılan
değerle aynı olmalı. Aksi halde kullanıcı ekranda gördüğüne güvenip
dosyada başka bir şey buluyor.
"""
from __future__ import annotations

import unittest

from core import ayar, betik, device_map, dogrulama, excel, isler, kontrol

from .ortak import ServisTesti
from . import sahte


class KontrolOnizleme(ServisTesti):

    def env(self):
        return device_map.yukle(1, ayar.KOK / "DeviceMap.json")

    def test_sablon_yapisi_dosyadan_gelir(self):
        y = kontrol.sablon_yapisi()
        dv = betik.device_verify()

        # Sütun başlıkları şablonun başlık satırıyla birebir aynı
        adlar = [s["ad"] for s in y["sutunlar"]]
        for baslik in dv.FILLABLE + [dv.COL_IP_TEMPLATE]:
            self.assertIn(baslik, adlar, baslik)

        # Grup başlıkları toplamı sütun sayısını vermeli
        self.assertEqual(sum(g["genislik"] for g in y["gruplar"]),
                         len(y["sutunlar"]))
        self.assertEqual([g["ad"] for g in y["gruplar"]],
                         ["FİZİKSEL KONTROL", "DOĞRULAMA KRİTERİ",
                          "YAZILIM KONTROL"])
        self.assertGreater(len(y["bolumler"]), 5)

    def test_onizleme_okunan_degeri_gosterir_okunmayani_bos_birakir(self):
        env = self.env()
        gor = isler.gorunum(1)
        anons = next(c for c in env.cihazlar if c.yontem == "http")
        s = dogrulama.basarili({"surum": "1.2.5", "seri": "SN-9",
                                "calisma": "01:00:00"}, "http")
        s.nesil = isler.sonraki_nesil()
        gor.yaz(anons.id, s)

        v = kontrol.onizleme(env, gor.hepsi())
        basliklar = [x["ad"] for x in v["sutunlar"]]
        i_surum = basliklar.index("Versiyon")
        i_baglanti = basliklar.index("Bağlantı Bilgisi")
        i_durum = basliklar.index("Durum Açıklaması")

        satir = self._satir_bul(v, anons.ip_sablonu)
        self.assertEqual(satir["hucreler"][i_surum]["deger"], "1.2.5")
        self.assertEqual(satir["hucreler"][i_baglanti]["deger"], anons.ip)
        self.assertEqual(satir["hucreler"][i_durum]["deger"], "Aktif")
        self.assertEqual(satir["durum"], dogrulama.YESIL)

        # Okunmamış bir cihazın aynı sütunları BOŞ kalmalı
        baska = next(c for c in env.cihazlar
                     if c.id != anons.id and gor.al(c.id) is None)
        bos = self._satir_bul(v, baska.ip_sablonu)
        self.assertEqual(bos["hucreler"][i_surum]["deger"], "")
        self.assertEqual(bos["hucreler"][i_baglanti]["deger"], "")
        self.assertEqual(bos["durum"], dogrulama.GRI)

    def test_na_hucreleri_sablondan_gelir_ve_okunmadidan_ayridir(self):
        env = self.env()
        v = kontrol.onizleme(env, {})
        basliklar = [x["ad"] for x in v["sutunlar"]]
        i_hoparlor = basliklar.index("Hoparlör Ses Seviyesi")

        # Switch'te ses seviyesi geçerli bir alan değil -> N/A
        sw = env.switchler()[0]
        self.assertTrue(
            self._satir_bul(v, sw.ip_sablonu)["hucreler"][i_hoparlor]["na"])

        # Intercom'da geçerli -> N/A değil, yalnızca okunmamış
        intercom = next(c for c in env.cihazlar
                        if (c.subtype or "") == "Intercom")
        hucre = self._satir_bul(v, intercom.ip_sablonu)["hucreler"][i_hoparlor]
        self.assertFalse(hucre["na"])
        self.assertEqual(hucre["deger"], "")

    def test_onizleme_ile_excel_ayni_degeri_yazar(self):
        """Ekranda görünen ile dosyaya yazılan aynı olmalı."""
        import tempfile
        from pathlib import Path
        import openpyxl

        env = self.env()
        gor = isler.gorunum(1)
        anons = next(c for c in env.cihazlar if c.yontem == "http")
        s = dogrulama.basarili({"surum": "9.9.9"}, "http")
        s.nesil = isler.sonraki_nesil()
        gor.yaz(anons.id, s)

        v = kontrol.onizleme(env, gor.hepsi())
        basliklar = [x["ad"] for x in v["sutunlar"]]
        ekran = self._satir_bul(v, anons.ip_sablonu)[
            "hucreler"][basliklar.index("Versiyon")]["deger"]

        cikti = Path(tempfile.mkdtemp(prefix="kontrol-test-")) / "c.xlsx"
        excel.uret(env, gor.hepsi(), cikti=cikti)
        dv = betik.device_verify()
        ws = openpyxl.load_workbook(cikti)["Kontrol Listesi"]
        sutun = {ws.cell(dv.HEADER_ROW, c).value: c
                 for c in range(1, ws.max_column + 1)
                 if ws.cell(dv.HEADER_ROW, c).value}
        dosyada = None
        for r in range(dv.HEADER_ROW + 1, ws.max_row + 1):
            if ws.cell(r, sutun[dv.COL_IP_TEMPLATE]).value == anons.ip_sablonu:
                dosyada = ws.cell(r, sutun["Versiyon"]).value
                break
        self.assertEqual(ekran, dosyada)
        self.assertEqual(ekran, "9.9.9")

    def test_api_ucu_calisir(self):
        taban = self.servis_ac()
        kod, y = self.cagir(taban, "/api/kontrol?set=3")
        self.assertEqual(kod, 200)
        self.assertEqual(y["setNo"], 3)
        self.assertTrue(y["cikti"].endswith("_set3.xlsx"))
        ilk = y["bolumler"][0]["satirlar"][0]
        self.assertIn(".3.", ilk["ip"])       # set numarası çözülmüş
        self.assertIn("n", ilk["ipSablonu"])  # şablon hâlâ şablon

    def _satir_bul(self, onizleme, sablon):
        for b in onizleme["bolumler"]:
            for s in b["satirlar"]:
                if s["ipSablonu"] == sablon:
                    return s
        self.fail(f"{sablon} önizlemede yok")


class CanliIslem(ServisTesti):
    """Tarama sürerken ana listede canlı durum."""

    def _harita(self, n=6):
        cihazlar = [{
            "Name": f"Intercom_{i}", "IP": "127.0.0.1", "IsActive": True,
            "Type": "Announcement", "SubType": "Intercom", "Port": str(10 + i),
            "PBXExtension": str(2000 + i), "Status": {},
        } for i in range(1, n + 1)]
        return sahte.device_map(cihazlar, switch_ip="127.0.0.1")

    def test_ilerleme_kuyrukta_durur_cihaz_listesinde_degil(self):
        """Adım adım durum yalnız işte; cihaz listesi son durumu gösterir."""
        self.kur_harita(self._harita(8))
        with sahte.sessiz() as sessiz:
            self.switch_portu(sessiz.port)
            ayar.ANONS_PORT = sessiz.port
            taban = self.servis_ac()

            kod, y = self.cagir(taban, "/api/tarama", {"set": 1})

            # Canlı ilerleme iş satırlarında
            kod, tam = self.cagir(taban, f"/api/is?id={y['id']}")
            durumlar = {s["durum"] for s in tam["satirlar"]}
            self.assertTrue(durumlar & {"bekliyor", "calisiyor"}, durumlar)

            # Cihaz listesinde tarama durumu alanı YOK
            kod, d = self.cagir(taban, "/api/durum?set=1")
            self.assertTrue(d["aktifTarama"])
            for c in d["cihazlar"]:
                self.assertNotIn("islem", c)

            self.cagir(taban, "/api/is/iptal", {"id": y["id"]})
            self.isi_bekle(isler.YONETICI.bul(y["id"]), sure=25)
            kod, d = self.cagir(taban, "/api/durum?set=1")
            self.assertFalse(d["aktifTarama"])

    def test_bilgi_satiri_sayaclara_girmez(self):
        """Telemetri gibi ilerleme satırları cihaz sayısını şişirmemeli."""
        is_ = isler.Is("tarama", "deneme", 1)

        class SahteCihaz:
            id, ad, ip, yontem = "d1", "Cihaz", "127.0.0.1", "http"

        is_.satir_kur(SahteCihaz())
        is_.ozel_satir("telemetri", "Canlı MQTT telemetrisi",
                       durum="tamam", not_="3 kayıt")

        self.assertEqual(is_.sayilar()["toplam"], 1)
        self.assertEqual(len(is_.satirlar()), 2)
        # İlerleme de yalnız cihaz satırlarından hesaplanır
        is_.satir_guncelle("d1", "tamam", "bitti")
        self.assertEqual(is_.ilerleme(), 1.0)


if __name__ == "__main__":
    unittest.main()
