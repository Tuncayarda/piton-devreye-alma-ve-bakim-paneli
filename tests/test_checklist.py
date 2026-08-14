#!/usr/bin/env python3
"""Checklist preview and the live job indicator.

This screen has to be a faithful preview of the template: columns, groups and
N/A cells come from the template, and the value on screen must equal the value
written to Excel. Otherwise the user trusts the screen and finds something
else in the file.
"""
from __future__ import annotations

import unittest

from panel import checklist, jobs, script_loader, settings
from panel.checklist import columns as cols
from panel.inventory import device_map
from panel.probe import result as probe_result
from panel import status

from .support import fakes
from .support.base import ServiceTest


class ChecklistPreview(ServiceTest):

    def inventory(self):
        return device_map.load(1, settings.ROOT / "DeviceMap.json")

    def test_template_layout_comes_from_the_file(self):
        layout = checklist.template_layout()
        verify = script_loader.device_verify()

        # Every column the script fills resolves to a stable id in the
        # template's header row.
        ids = [column["id"] for column in layout["columns"]]
        for column in verify.FILLABLE + [verify.COL_IP_TEMPLATE]:
            self.assertIn(column, ids, column)
        # ...and each of those ids carries its documented heading.
        for column in layout["columns"]:
            if column["id"]:
                self.assertEqual(column["name"],
                                 cols.HEADING_FOR_COLUMN[column["id"]])

        # The group widths must add up to the column count
        self.assertEqual(sum(g["width"] for g in layout["groups"]),
                         len(layout["columns"]))
        self.assertEqual([g["name"] for g in layout["groups"]],
                         ["PHYSICAL CHECK", "VERIFICATION CRITERIA",
                          "SOFTWARE CHECK"])
        self.assertGreater(len(layout["sections"]), 5)

    def test_preview_shows_read_values_and_leaves_unread_empty(self):
        inventory = self.inventory()
        view = jobs.view_for(1)
        device = next(d for d in inventory.devices if d.read_method == "http")
        result = probe_result.success(
            {"version": "1.2.5", "serial": "SN-9", "uptime": "01:00:00"},
            "http")
        result.generation = jobs.next_generation()
        view.write(device.id, result)

        preview = checklist.build_preview(inventory, view.all())
        ids = [c["id"] for c in preview["columns"]]
        i_version = ids.index(cols.VERSION)
        i_link = ids.index(cols.CONNECTION_INFO)
        i_state = ids.index(cols.STATUS_DESCRIPTION)

        row = self._find_row(preview, device.ip_template)
        self.assertEqual(row["cells"][i_version]["value"], "1.2.5")
        self.assertEqual(row["cells"][i_link]["value"], device.ip)
        self.assertEqual(row["cells"][i_state]["value"], cols.STATUS_ACTIVE)
        self.assertEqual(row["state"], status.OK)

        # The same columns of an unread device must stay EMPTY
        other = next(d for d in inventory.devices
                     if d.id != device.id and view.get(d.id) is None)
        empty = self._find_row(preview, other.ip_template)
        self.assertEqual(empty["cells"][i_version]["value"], "")
        self.assertEqual(empty["cells"][i_link]["value"], "")
        self.assertEqual(empty["state"], status.UNKNOWN)

    def test_na_cells_come_from_the_template_and_differ_from_unread(self):
        inventory = self.inventory()
        preview = checklist.build_preview(inventory, {})
        ids = [c["id"] for c in preview["columns"]]
        i_speaker = ids.index(cols.SPEAKER_VOLUME)

        # Volume is not a valid field on a switch -> N/A
        switch = inventory.switches()[0]
        self.assertTrue(self._find_row(preview, switch.ip_template)
                        ["cells"][i_speaker]["notApplicable"])

        # Valid on an Intercom -> not N/A, merely unread
        intercom = next(d for d in inventory.devices
                        if (d.subtype or "") == "Intercom")
        cell = self._find_row(preview,
                              intercom.ip_template)["cells"][i_speaker]
        self.assertFalse(cell["notApplicable"])
        self.assertEqual(cell["value"], "")

    def test_preview_and_excel_write_the_same_value(self):
        """What is on screen must equal what is written to the file."""
        import tempfile
        from pathlib import Path
        import openpyxl

        inventory = self.inventory()
        view = jobs.view_for(1)
        device = next(d for d in inventory.devices if d.read_method == "http")
        result = probe_result.success({"version": "9.9.9"}, "http")
        result.generation = jobs.next_generation()
        view.write(device.id, result)

        preview = checklist.build_preview(inventory, view.all())
        ids = [c["id"] for c in preview["columns"]]
        on_screen = self._find_row(preview, device.ip_template)[
            "cells"][ids.index(cols.VERSION)]["value"]

        output = Path(tempfile.mkdtemp(prefix="checklist-test-")) / "c.xlsx"
        checklist.export(inventory, view.all(), output=output)
        verify = script_loader.device_verify()
        sheet = openpyxl.load_workbook(output)[checklist.workbook.SHEET_NAME]
        columns = {}
        for index in range(1, sheet.max_column + 1):
            column = cols.column_for_heading(
                sheet.cell(verify.HEADER_ROW, index).value)
            if column:
                columns[column] = index
        in_file = None
        for r in range(verify.HEADER_ROW + 1, sheet.max_row + 1):
            cell = sheet.cell(r, columns[verify.COL_IP_TEMPLATE]).value
            if cell == device.ip_template:
                in_file = sheet.cell(r, columns[cols.VERSION]).value
                break
        self.assertEqual(on_screen, in_file)
        self.assertEqual(on_screen, "9.9.9")

    def test_api_endpoint_works(self):
        base = self.start_service()
        code, body = self.call(base, "/api/checklist?set=3")
        self.assertEqual(code, 200)
        self.assertEqual(body["setNo"], 3)
        self.assertTrue(body["output"].endswith("_set3.xlsx"))
        first = body["sections"][0]["rows"][0]
        self.assertIn(".3.", first["ip"])        # set number resolved
        self.assertIn("n", first["ipTemplate"])  # the template stays a template

    def _find_row(self, preview, template):
        for section in preview["sections"]:
            for row in section["rows"]:
                if row["ipTemplate"] == template:
                    return row
        self.fail(f"{template} is not in the preview")


class LiveJob(ServiceTest):
    """Live state in the main list while a scan is running."""

    def _topology(self, n=6):
        devices = [{
            "Name": f"Intercom_{i}", "IP": "127.0.0.1", "IsActive": True,
            "Type": "Announcement", "SubType": "Intercom", "Port": str(10 + i),
            "PBXExtension": str(2000 + i), "Status": {},
        } for i in range(1, n + 1)]
        return fakes.device_map(devices, switch_ip="127.0.0.1")

    def test_progress_lives_in_the_queue_not_the_device_list(self):
        """Step-by-step state belongs to the job; the list shows last state."""
        self.build_map(self._topology(8))
        with fakes.silent() as silent:
            self.switch_port(silent.port)
            settings.ANNOUNCEMENT_PORT = silent.port
            base = self.start_service()

            code, started = self.call(base, "/api/scan", {"set": 1})

            # Live progress sits in the job's rows
            code, full = self.call(base, f"/api/job?id={started['id']}")
            states = {row["state"] for row in full["rows"]}
            self.assertTrue(states & {"queued", "running"}, states)

            # The device list carries NO scan-progress field
            code, state = self.call(base, "/api/state?set=1")
            self.assertTrue(state["scanRunning"])
            for device in state["devices"]:
                self.assertNotIn("job", device)

            self.call(base, "/api/job/cancel", {"id": started["id"]})
            self.await_job(jobs.QUEUE.find(started["id"]), timeout=25)
            code, state = self.call(base, "/api/state?set=1")
            self.assertFalse(state["scanRunning"])

    def test_informational_rows_stay_out_of_the_counters(self):
        """Progress rows like telemetry must not inflate the device count."""
        job = jobs.Job("scan", "trial", 1)

        class FakeDevice:
            id, name, ip, read_method = "d1", "Device", "127.0.0.1", "http"

        job.add_device_row(FakeDevice())
        job.add_row("telemetry", "Live MQTT telemetry",
                    state="done", note="3 records")

        self.assertEqual(job.counts()["total"], 1)
        self.assertEqual(len(job.rows()), 2)
        # Progress is computed from device rows only
        job.update_row("d1", "done", "finished")
        self.assertEqual(job.progress(), 1.0)


if __name__ == "__main__":
    unittest.main()
