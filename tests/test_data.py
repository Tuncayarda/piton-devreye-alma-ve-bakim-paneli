#!/usr/bin/env python3
"""Data presentation, inventory and the Excel export.

One concern runs through these tests: the panel must never show data that
exists nowhere as though it did. "Not read" and "not applicable" are separate,
there are no fake defaults, and only a genuinely read value is written to
Excel.
"""
from __future__ import annotations

import ipaddress
import json
import tempfile
import time
import unittest
from unittest import mock
from pathlib import Path
from typing import ClassVar

from panel import (checklist, editions, ip_assign, jobs, script_loader,
                   settings, status)
from panel.checklist import columns as cols
from panel.inventory import catalog, device_map
from panel.probe import android, reader
from panel.probe import result as probe_result
from panel.telemetry import TelemetrySnapshot

from .support import fakes
from .support.base import YATAKLI_MAP, PanelTest, ServiceTest


class Inventory(PanelTest):

    def test_the_n_in_an_ip_template_is_resolved(self):
        self.assertEqual(device_map.resolve_template("10.n.1.24", 3),
                         "10.3.1.24")
        self.assertEqual(device_map.resolve_template("10.n.1.24", 12),
                         "10.12.1.24")
        # An 'n' inside another word is left alone
        self.assertEqual(device_map.resolve_template("10.no.1.24", 3),
                         "10.no.1.24")
        self.assertEqual(device_map.to_template("10.3.1.24", 3), "10.n.1.24")

    def test_the_real_devicemap_resolves_for_every_set(self):
        for n in (1, 2, 9, 16):
            inventory = device_map.load(n, YATAKLI_MAP)
            self.assertEqual(inventory.set_no, n)
            self.assertGreater(len(inventory.devices), 30)
            for device in inventory.devices:
                self.assertNotIn(".n.", device.ip,
                                 f"the {device.name} template was not resolved")
                self.assertTrue(device.ip.startswith(f"10.{n}."), device.ip)

    def test_the_dto_carries_no_credential_fields(self):
        inventory = device_map.load(1, YATAKLI_MAP)
        for device in inventory.devices:
            data = device.dto()
            self.assertNotIn("Password", str(data))
            self.assertNotIn("Username", str(data))
            self.assertNotIn("password", str(data).lower())
            for name in ("id", "name", "ip", "ipTemplate", "type", "subtype",
                         "switch", "port", "category", "readMethod"):
                self.assertIn(name, data, name)
            self.assertNotIn("Password", device.extra)
            self.assertNotIn("Username", device.extra)

    def test_device_ids_are_unique_and_stable(self):
        a = device_map.load(1, YATAKLI_MAP)
        b = device_map.load(4, YATAKLI_MAP)
        ids = [device.id for device in a.devices]
        self.assertEqual(len(ids), len(set(ids)))
        # The id is independent of the set number; the IP changes
        self.assertEqual(ids, [device.id for device in b.devices])
        self.assertNotEqual([d.ip for d in a.devices],
                            [d.ip for d in b.devices])


class DataPresentation(PanelTest):

    def test_not_read_and_not_applicable_are_separate(self):
        not_read = probe_result.not_read("mqtt")
        not_applicable = probe_result.not_applicable(
            "mqtt", "Not applicable on this device")
        self.assertEqual(not_read.state, status.UNKNOWN)
        self.assertEqual(not_applicable.state, status.UNKNOWN)
        self.assertNotEqual(not_read.verification, not_applicable.verification)
        self.assertEqual(not_read.verification, status.NOT_READ)
        self.assertEqual(not_applicable.verification, status.NOT_APPLICABLE)

    def test_no_fake_defaults_for_unread_fields(self):
        """If the device does not answer, no field is invented."""
        topology = fakes.device_map([], switch_ip="127.0.0.1")
        inventory = self.build_map(topology)
        with fakes.silent() as silent:
            self.switch_port(silent.port)
            result = reader.read_device(inventory.switches()[0],
                                        credentials=("a", "b"), timeout=1.0)
        self.assertEqual(result.state, status.FAILED)
        self.assertEqual(result.fields, {})

    def test_without_mqtt_a_device_stays_grey_not_failed(self):
        """With no broker an MQTT device is 'not read', not 'failed'."""
        topology = fakes.device_map([{
            "Name": "Ap_1", "IP": "127.0.0.1", "IsActive": True,
            "Type": "AP", "SubType": "", "Port": "26", "Status": {},
        }])
        inventory = self.build_map(topology)
        ap = inventory.by_type("AP")[0]
        result = reader.read_device(ap, telemetry=None)
        self.assertEqual(result.state, status.UNKNOWN)
        self.assertEqual(result.verification, status.NOT_READ)
        self.assertIn("MQTT", result.detail)

    def test_the_counters_are_per_state(self):
        counts = probe_result.tally([
            probe_result.success({}, "http"),
            probe_result.success({}, "http"),
            probe_result.ProbeResult(state=status.AUTH),
            probe_result.ProbeResult(state=status.FAILED),
            probe_result.not_read("mqtt"),
        ])
        self.assertEqual(counts, {"ok": 2, "auth": 1, "failed": 1,
                                  "unknown": 1})

    def test_uptime_text(self):
        self.assertEqual(probe_result.uptime_text(3661), "01:01:01")
        self.assertEqual(probe_result.uptime_text(-1), "")
        self.assertEqual(probe_result.uptime_text(None), "")
        self.assertEqual(probe_result.uptime_text("abc"), "")


class AnnouncementFields(PanelTest):
    """Gain and the outbound number are separate fields, not the volume."""

    def build(self):
        topology = fakes.device_map([{
            "Name": "Intercom_1", "IP": "127.0.0.1", "IsActive": True,
            "Type": "Announcement", "SubType": "Intercom", "Port": "11",
            "PBXExtension": "2001", "Status": {"NoError": True},
        }])
        return self.build_map(topology).by_type("Announcement")[0]

    def test_gain_and_outbound_number_are_read(self):
        device = self.build()
        with fakes.announcement() as fake:
            settings.ANNOUNCEMENT_PORT = fake.port
            result = reader.read_device(device)
        f = result.fields
        self.assertEqual(result.state, status.OK)
        self.assertEqual(f["speakerGain"], 4)
        self.assertEqual(f["micGain"], 2)
        self.assertEqual(f["sipOutbound"], "5001")
        # Gain does not stand in for the volume
        self.assertEqual(f["speakerVolume"], 70)
        self.assertEqual(f["micVolume"], 60)
        # The outbound number is not confused with the device's own extension
        self.assertEqual(f["sipExtension"], "2001")

    def test_a_missing_field_is_not_invented(self):
        device = self.build()
        partial = {"firmwareVersion": "1.2.5", "uptime": 10,
                   "pbxExtension": "2001"}
        with fakes.announcement(settings=partial) as fake:
            settings.ANNOUNCEMENT_PORT = fake.port
            f = reader.read_device(device).fields
        self.assertIsNone(f["speakerGain"])
        self.assertIsNone(f["micGain"])
        self.assertEqual(f["sipOutbound"], "")

    def test_the_excel_mapping_covers_the_new_columns(self):
        for name, column in (("speakerGain", cols.SPEAKER_GAIN),
                             ("micGain", cols.MIC_GAIN),
                             ("sipOutbound", cols.SIP_OUTBOUND)):
            self.assertEqual(checklist.COLUMN_FOR_FIELD.get(name), column,
                             name)


class TelemetryMatching(PanelTest):
    """The broker publishes template IPs, the panel looks up resolved ones.

    When the two did not match, every MQTT-backed device dropped to red with
    "not found in telemetry".
    """

    def build_telemetry(self, set_no=1):
        snapshot = TelemetrySnapshot("10.1.1.1")
        snapshot.records.clear()
        snapshot._index("10.n.1.4", {"Name": "Hmi", "IP": "10.n.1.4",
                                     "Status": {"NoError": True,
                                                "Uptime": 27241}},
                        set_no)
        snapshot.apps = {"ClientManager_MCP_YATAKLI_1": {
            "ClientId": "ClientManager_MCP_YATAKLI_1", "DeviceIP": "10.1.1.4",
            "HWID": "34DA8534", "Status": "connected", "Version": "1.2.5"}}
        return snapshot

    def test_a_template_ip_is_found_by_its_resolved_ip(self):
        snapshot = self.build_telemetry()
        self.assertIsNotNone(snapshot.record("10.1.1.4"))    # resolved
        self.assertIsNotNone(snapshot.record("10.n.1.4"))    # template
        self.assertIsNone(snapshot.record("10.1.1.9"))

    def test_the_hmi_uptime_is_completed_from_devicemap(self):
        """The AppStatus payload has no Uptime; it comes from the record."""
        topology = fakes.device_map([{
            "Name": "Hmi", "IP": "10.n.1.4", "IsActive": True,
            "Type": "HMI", "SubType": "", "Port": "6", "Status": {},
        }], switch_ip="10.n.1.100")
        inventory = self.build_map(topology)
        hmi = inventory.by_type("HMI")[0]
        result = reader.read_device(hmi, telemetry=self.build_telemetry())
        self.assertEqual(result.state, status.OK)
        self.assertEqual(result.fields["version"], "1.2.5")
        self.assertEqual(result.fields["uptime"], "07:34:01")

    def test_an_uptime_of_minus_one_is_not_an_uptime(self):
        """PISCU reports -1 for a device that is off; that is not a duration."""
        topology = fakes.device_map([{
            "Name": "Hmi", "IP": "10.n.1.4", "IsActive": True,
            "Type": "HMI", "SubType": "", "Port": "6", "Status": {},
        }], switch_ip="10.n.1.100")
        inventory = self.build_map(topology)
        snapshot = self.build_telemetry()
        snapshot.records["10.1.1.4"]["Status"]["Uptime"] = -1
        result = reader.read_device(inventory.by_type("HMI")[0],
                                    telemetry=snapshot)
        self.assertEqual(result.fields["uptime"], "")


class RetainedMessageIsNotDevicePresence(PanelTest):
    """A retained broker message stays there after the device is gone.

    Seen in the field: with the HMI's cable unplugged the panel showed it green
    and "Verified". The device's IP never answered; the panel read the message
    left in the broker and thought it was up. The note it showed even said
    "disconnected".

    The payloads here are copied verbatim from the field broker.
    """

    HMI: ClassVar[dict] = {
        "Name": "Hmi", "IP": "10.n.1.4", "IsActive": True,
        "Type": "HMI", "SubType": "", "Port": "6", "Status": {},
    }
    ICU: ClassVar[dict] = {
        "Name": "Icu", "IP": "10.n.1.2", "IsActive": True,
        "Type": "ICU", "SubType": "", "Port": "5", "Status": {},
    }

    def _read(self, live, apps, definition=None):
        """Reads a single device against the given broker picture.

        `live`  the ALFA/DeviceMap record (None: no record at all)
        `apps`  the ALFA/AppStatus messages {ClientId: payload}
        """
        definition = definition or self.HMI
        inventory = self.build_map(
            fakes.device_map([definition], switch_ip="10.n.1.100"))
        snapshot = TelemetrySnapshot("10.1.1.1")
        snapshot.records.clear()
        if live is not None:
            snapshot._index(definition["IP"], live, 1)
        snapshot.apps = apps
        return reader.read_device(
            inventory.by_type(definition["Type"])[0], telemetry=snapshot)

    # ── AppStatus: last will (LWT) ──
    def test_a_disconnected_lwt_does_not_turn_green(self):
        """The field payload: {"ClientId": "...MCP...", "Status":
        "disconnected"}

        None of DeviceIP, HWID or Version is present. The panel found the
        record by ClientId and called it "verified" with empty fields.
        """
        result = self._read(
            {"IP": "10.n.1.4", "Status": {"NoError": False, "Uptime": -1}},
            {"ClientManager_MCP_YATAKLI_1": {
                "ClientId": "ClientManager_MCP_YATAKLI_1",
                "Status": "disconnected"}})
        self.assertEqual(result.state, status.FAILED)
        self.assertIn("not connected", result.detail)
        self.assertIn("disconnected", result.detail)

    def test_a_connected_device_stays_green(self):
        """The fix must not break a working device."""
        result = self._read(
            {"IP": "10.n.1.4", "Status": {"NoError": True, "Uptime": 27241}},
            {"ClientManager_MCP_YATAKLI_1": {
                "ClientId": "ClientManager_MCP_YATAKLI_1",
                "DeviceIP": "10.1.1.4", "HWID": "34DA8534",
                "Status": "connected", "Version": "1.2.5"}})
        self.assertEqual(result.state, status.OK)
        self.assertEqual(result.fields["version"], "1.2.5")

    def test_appstatus_connected_loses_to_a_faulty_live_record(self):
        """When two signals disagree, the device does not count as up.

        The AppStatus message may still read "connected" after the device is
        gone (if power is cut before the will can be published). PISCU's live
        record, however, is watching the device right now.
        """
        result = self._read(
            {"IP": "10.n.1.4", "Status": {"NoError": False, "Uptime": -1}},
            {"ClientManager_MCP_YATAKLI_1": {
                "ClientId": "ClientManager_MCP_YATAKLI_1",
                "DeviceIP": "10.1.1.4", "HWID": "34DA8534",
                "Status": "connected", "Version": "1.2.5"}})
        self.assertEqual(result.state, status.FAILED)
        self.assertEqual(
            result.detail,
            "PISCU reports the device as powered down or faulty")

    def test_an_old_payload_without_status_is_judged_by_the_live_record(self):
        """With no Status field in AppStatus, the live record is the only
        signal."""
        result = self._read(
            {"IP": "10.n.1.4", "Status": {"NoError": True, "Uptime": 100}},
            {"ClientManager_MCP_YATAKLI_1": {
                "ClientId": "ClientManager_MCP_YATAKLI_1",
                "DeviceIP": "10.1.1.4", "Version": "1.2.5"}})
        self.assertEqual(result.state, status.OK)

    # ── devices backed by the live DeviceMap (ICU, AP, LED, Landing LCD) ──
    def test_a_faulty_device_is_an_error_not_not_applicable(self):
        """A device that is off used to show as "N/A" (grey).

        Grey means "this check does not exist on this device". Here the check
        WAS made and the device reported faulty — that is a result, not an
        absence.
        """
        result = self._read(
            {"IP": "10.n.1.2", "Status": {"NoError": False, "Uptime": -1}},
            {}, definition=self.ICU)
        self.assertEqual(result.state, status.FAILED)
        self.assertEqual(result.verification, status.UNVERIFIED)
        self.assertEqual(
            result.detail,
            "PISCU reports the device as powered down or faulty")

    def test_a_healthy_mqtt_device_stays_green(self):
        result = self._read(
            {"IP": "10.n.1.2", "SerialNumber": "ICU-9",
             "Status": {"NoError": True, "Uptime": 500, "Version": "2.0"}},
            {}, definition=self.ICU)
        self.assertEqual(result.state, status.OK)
        self.assertEqual(result.fields["serial"], "ICU-9")


# ── Compartment LCD / ADB ───────────────────────────────────────────────
DUMPSYS = """Packages:
  Package [com.piton.train_lcd_panel] (5b3a8c0):
    userId=10123
    codePath=/data/app/com.piton.train_lcd_panel-1
    versionCode=1 minSdk=21 targetSdk=35
    versionName=0.0.5
    flags=[ HAS_CODE ALLOW_CLEAR_USER_DATA ]
    firstInstallTime=2026-07-07 13:13:40
    lastUpdateTime=2026-07-28 08:05:32
"""

LOGCAT = """08-05 12:04:17.249 13854 13946 I AnnounceSip: Registration state=registered code=200
08-05 12:24:25.747 13465 13557 I AnnounceSip: SIP engine started: sip:6001@10.1.1.1:5060 (UDP)
08-05 12:24:25.757 13465 13557 I AnnounceSip: Registration state=registered code=200
"""


class FakeAdb:
    """Stands in for subprocess.run; returns canned output for adb calls."""

    def __init__(self, **answers):
        self.answers = answers
        self.calls: list[list[str]] = []

    def __call__(self, args, **kwargs):
        self.calls.append(list(args))
        joined = " ".join(args)
        output = ""
        if "connect" in args:
            output = "connected"
        elif "ro.serialno" in joined:
            output = self.answers.get("serial", "rk3568r0001")
        elif "persist.sys.timezone" in joined:
            output = self.answers.get("tz", "CST-3:00:00")
        elif "/proc/uptime" in joined:
            output = self.answers.get("uptime", "3661.20 12000.00")
        elif "dumpsys" in args:
            output = self.answers.get("dumpsys", DUMPSYS)
        elif "logcat" in args:
            output = self.answers.get("logcat", LOGCAT)

        class Result:
            stdout = output
            stderr = ""
            returncode = 0
        return Result()


class CompartmentLcd(PanelTest):
    """The ADB read: connecting is not enough, data must arrive too.

    Showing a device green when adb connects but the app version cannot be
    read reported an uninstalled app as installed in the field.
    """

    def build_lcd(self):
        topology = fakes.device_map([{
            "Name": "Compartment_Lcd_1", "IP": "10.n.1.40", "IsActive": True,
            "Type": "LCD", "SubType": "Compartment", "Port": "13",
            "PBXExtension": "6001", "Status": {},
        }])
        inventory = self.build_map(topology)
        return inventory.by_type("LCD")[0]

    def patch(self, fake_adb):
        previous = android.subprocess.run
        android.subprocess.run = fake_adb
        self.addCleanup(
            lambda: setattr(android.subprocess, "run", previous))

    def test_the_dumpsys_version_is_parsed(self):
        info = android.package_info(DUMPSYS)
        self.assertEqual(info["version"], "0.0.5")
        self.assertEqual(info["versionCode"], "1")
        self.assertEqual(info["minSdk"], "21")
        self.assertEqual(info["targetSdk"], "35")
        self.assertEqual(info["installedAt"], "2026-07-07 13:13:40")
        self.assertEqual(info["updatedAt"], "2026-07-28 08:05:32")

    def test_the_version_is_empty_when_the_package_is_missing(self):
        info = android.package_info(
            "Unable to find package: com.piton.train_lcd_panel")
        self.assertEqual(info["version"], "")

    def test_the_extension_and_registration_come_from_the_sip_log(self):
        data = android.sip_log(LOGCAT)
        self.assertEqual(data["sipExtension"], "6001")
        self.assertEqual(data["sipPbx"], "10.1.1.1")
        self.assertEqual(data["sipPort"], "5060")
        self.assertEqual(data["sipTransport"], "UDP")
        self.assertEqual(data["sipRegistration"], "registered")
        self.assertEqual(data["sipCode"], "200")

    def test_the_fields_stay_empty_without_a_sip_line(self):
        data = android.sip_log("")
        self.assertEqual(data["sipExtension"], "")
        self.assertEqual(data["sipRegistration"], "")

    def test_a_read_lcd_is_green_with_its_fields_filled(self):
        device = self.build_lcd()
        self.patch(FakeAdb())
        result = reader.read_device(device)
        self.assertEqual(result.state, status.OK)
        self.assertEqual(result.fields["version"], "0.0.5")
        self.assertEqual(result.fields["serial"], "rk3568r0001")
        self.assertEqual(result.fields["sipExtension"], "6001")
        self.assertEqual(result.fields["sipPbx"], "10.1.1.1")
        self.assertEqual(result.fields["sipRegistration"], "registered (200)")
        self.assertEqual(result.fields["uptime"], "01:01:01")
        self.assertEqual(result.fields["timezone"], "CST-3:00:00")

    def test_an_lcd_whose_version_cannot_be_read_is_not_green(self):
        """The original complaint: it showed 'Active' with no data at all."""
        device = self.build_lcd()
        self.patch(FakeAdb(dumpsys="Unable to find package: com.piton.x"))
        result = reader.read_device(device)
        self.assertEqual(result.state, status.FAILED)
        self.assertEqual(result.verification, status.UNVERIFIED)
        self.assertEqual(result.fields, {})
        self.assertIn("Could not read", result.detail)

    def test_a_display_without_the_app_is_green_on_a_stand(self):
        """The demonstration case, and THE PROJECT is what decides it.

        A stand carries borrowed hardware running whatever each unit happens
        to have on it. Demanding one named application turned the whole board
        red and hid the units that genuinely could not be reached. Relaxed,
        the display still has to ANSWER — what is dropped is only the demand
        for that application.
        """
        device = self.build_lcd()
        self.patch(FakeAdb(dumpsys="Unable to find package: com.piton.x"))
        with mock.patch.object(editions, "on_a_stand", return_value=True):
            result = reader.read_device(device)
        self.assertEqual(result.state, status.OK)
        # The connection's own findings are real and are kept.
        self.assertEqual(result.fields["serial"], "rk3568r0001")
        self.assertEqual(result.fields["uptime"], "01:01:01")
        # ...and the version is honestly empty rather than invented.
        self.assertEqual(result.fields["version"], "")
        # The row does not claim to have read an application it did not find.
        self.assertNotIn("0.0.5", result.detail)
        self.assertIn("rk3568r0001", result.detail)

    def test_a_silent_display_is_still_red_on_a_stand(self):
        """Relaxing the check must not turn "unreachable" into "fine"."""
        device = self.build_lcd()
        self.patch(FakeAdb(serial=""))
        with mock.patch.object(editions, "on_a_stand", return_value=True):
            result = reader.read_device(device)
        self.assertEqual(result.state, status.FAILED)

    def test_a_train_still_demands_the_app(self):
        """The default, and it is not a setting anybody has to remember."""
        device = self.build_lcd()
        self.patch(FakeAdb(dumpsys="Unable to find package: com.piton.x"))
        with mock.patch.object(editions, "on_a_stand", return_value=False):
            result = reader.read_device(device)
        self.assertEqual(result.state, status.FAILED)

    def test_only_the_stand_project_carries_the_flag(self):
        """It is on the FUAR project and on nothing else."""
        from panel.editions import catalogue

        stands = [p.key for p in catalogue.ALL_PROJECTS if p.stand]
        self.assertEqual(stands, ["fuar"])

    def test_the_setting_overrides_the_project_in_both_directions(self):
        """For a bench that is neither a train nor a stand."""
        device = self.build_lcd()
        self.patch(FakeAdb(dumpsys="Unable to find package: com.piton.x"))
        with mock.patch.object(editions, "on_a_stand", return_value=False), \
                mock.patch.object(settings, "ADB_REQUIRE_PACKAGE", "0"):
            self.assertEqual(reader.read_device(device).state, status.OK)
        with mock.patch.object(editions, "on_a_stand", return_value=True), \
                mock.patch.object(settings, "ADB_REQUIRE_PACKAGE", "1"):
            self.assertEqual(reader.read_device(device).state, status.FAILED)

    def test_an_unreachable_lcd_is_red(self):
        device = self.build_lcd()
        self.patch(FakeAdb(serial=""))
        result = reader.read_device(device)
        self.assertEqual(result.state, status.FAILED)
        self.assertIn("adb", result.detail)

    def test_the_read_is_still_valid_without_a_sip_registration(self):
        """The log buffer may have wrapped; a missing SIP line is not an error."""
        device = self.build_lcd()
        self.patch(FakeAdb(logcat=""))
        result = reader.read_device(device)
        self.assertEqual(result.state, status.OK)
        self.assertEqual(result.fields["version"], "0.0.5")
        self.assertEqual(result.fields["sipExtension"], "")

    def test_the_extension_falls_back_to_the_broker_announcement(self):
        """The extension must be readable without restarting the app."""
        device = self.build_lcd()
        snapshot = TelemetrySnapshot("10.1.1.1")
        snapshot.sip = {device.ip: "6001"}

        self.patch(FakeAdb(logcat=""))
        result = reader.read_device(device, telemetry=snapshot)
        self.assertEqual(result.fields["sipExtension"], "6001")
        self.assertIn("ALFA/SipPort", result.fields["sipExtensionSource"])

        # When the log has it, the device's own statement wins
        self.patch(FakeAdb())
        result = reader.read_device(device, telemetry=snapshot)
        self.assertEqual(result.fields["sipExtension"], "6001")
        self.assertEqual(result.fields["sipExtensionSource"], "device log")

    def test_a_registered_pbx_address_is_completed_from_the_project(self):
        """With the log's `sip:...@10.1.1.1` line gone, the PBX is the set's
        PISCU."""
        device = self.build_lcd()
        self.patch(FakeAdb(logcat=(
            "08-05 12:04:17.249 I AnnounceSip: "
            "Registration state=registered code=200\n")))
        f = reader.read_device(device, pbx_ip="10.1.1.1").fields
        self.assertEqual(f["sipPbx"], "10.1.1.1")
        self.assertEqual(f["sipPbxSource"], "project (PISCU)")

        # When the log has it, the device's own statement wins
        self.patch(FakeAdb())
        f = reader.read_device(device, pbx_ip="10.9.9.9").fields
        self.assertEqual(f["sipPbx"], "10.1.1.1")
        self.assertEqual(f["sipPbxSource"], "device log")

    def test_no_pbx_is_written_when_the_device_is_not_registered(self):
        """Writing a PBX for an unregistered device shows a link that is not
        there."""
        device = self.build_lcd()
        self.patch(FakeAdb(logcat=(
            "08-05 12:04:17.249 I AnnounceSip: "
            "Registration state=failed code=408\n")))
        f = reader.read_device(device, pbx_ip="10.1.1.1").fields
        self.assertEqual(f["sipPbx"], "")
        self.assertEqual(f["sipPbxSource"], "")
        self.assertEqual(f["sipRegistration"], "failed (408)")

    def test_no_number_is_invented_without_a_broker_announcement(self):
        device = self.build_lcd()
        snapshot = TelemetrySnapshot("10.1.1.1")
        snapshot.sip = {"10.9.9.9": "6001"}      # another device's number
        self.patch(FakeAdb(logcat=""))
        result = reader.read_device(device, telemetry=snapshot)
        self.assertEqual(result.fields["sipExtension"], "")
        self.assertEqual(result.fields["sipExtensionSource"], "")


class IpPlan(PanelTest):

    def test_ip_assignment_accepts_only_groups_with_a_commissioning_runner(self):
        supported = [g["name"] for g in catalog.GROUPS
                     if catalog.group_supports(g, "ip")]
        self.assertEqual(supported, ["Intercom", "Compartment LCD"])
        self.assertEqual(
            [g["name"] for g in ip_assign.resolve_groups(
                ["Intercom", "Compartment LCD", "Handset", "Camera",
                 "All"])],
            ["Intercom", "Compartment LCD"])

    def test_port_parsing(self):
        allowed = set(range(1, 25))
        self.assertEqual(ip_assign.parse_ports("11-14", allowed),
                         [11, 12, 13, 14])
        self.assertEqual(ip_assign.parse_ports("11-13, 21 22", allowed),
                         [11, 12, 13, 21, 22])
        self.assertEqual(ip_assign.format_ports([11, 12, 13, 21]), "11-13, 21")
        for bad in ("14-11", "abc", "1-", "-", "11;;;x"):
            with self.assertRaises(ValueError, msg=bad):
                ip_assign.parse_ports(bad, allowed)

    def test_a_port_outside_the_allowed_set_is_rejected(self):
        with self.assertRaises(ValueError) as caught:
            ip_assign.parse_ports("11-14", {11, 12})
        self.assertIn("not defined on this switch", str(caught.exception))

    def test_the_factory_ip_does_not_change_with_the_set(self):
        """A device out of the box does not know which set it will join.

        Resolving the template would search 10.8.1.12 on set 8 and never find
        anything.
        """
        topology = fakes.device_map([
            {"Name": "Intercom_1", "IP": "10.n.1.10", "IsActive": True,
             "Type": "Announcement", "SubType": "Intercom", "Port": "11",
             "Status": {}}], switch_ip="10.n.1.101")
        self.build_map(topology)
        for n in (1, 8, 112):
            inventory = device_map.load(n, self.map_path)
            self.assertEqual(ip_assign.factory_ip(inventory), "10.1.1.12")
            plan = ip_assign.build_plan(inventory, "Intercom", [11])
            self.assertEqual(plan["rows"][0]["factoryIp"], "10.1.1.12")
            # The target IP still resolves per set; only the factory is fixed.
            self.assertEqual(plan["rows"][0]["targetIp"], f"10.{n}.1.10")

    def test_search_candidates_come_from_the_network_and_mask(self):
        candidates = ip_assign.search_candidates("10.1.1.0", "255.255.255.0")
        self.assertEqual(candidates[0], "10.1.1.1")
        self.assertEqual(candidates[-1], "10.1.1.254")
        with self.assertRaises(ValueError) as caught:
            ip_assign.search_candidates("10.0.0.0", "255.0.0.0")
        self.assertIn("too wide", str(caught.exception))

    def test_a_network_too_wide_is_refused_without_being_built(self):
        """The limit is decided by counting, never by building the list.

        A /8 holds 16.7 million addresses. Building them to find out there
        are too many took eighteen seconds and about a gigabyte — on the
        request thread, in an elevated process, for a mask the user can
        type by accident. The refusal must cost nothing.
        """
        started = time.monotonic()
        with self.assertRaises(ValueError) as caught:
            ip_assign.search_candidates("10.0.0.0", "255.0.0.0")
        self.assertLess(time.monotonic() - started, 1.0,
                        "the address list was built before it was refused")
        # The count in the message is the usable host count, as before.
        self.assertIn("16777214", str(caught.exception))

    def test_the_refused_count_matches_what_would_have_been_built(self):
        """Counting and building must not disagree at the edges.

        `hosts()` treats /31 and /32 specially, so a count derived from
        `num_addresses` has to as well — otherwise a mask one bit either
        side of the limit is refused or accepted wrongly.
        """
        for prefix in range(23, 33):
            network = ipaddress.ip_network(f"10.1.1.0/{prefix}", strict=False)
            built = len(list(network.hosts())) or 1
            candidates = ip_assign.search_candidates(
                "10.1.1.0", str(network.netmask), limit=built)
            self.assertEqual(len(candidates), built, f"/{prefix}")
            with self.assertRaises(ValueError, msg=f"/{prefix}"):
                ip_assign.search_candidates(
                    "10.1.1.0", str(network.netmask), limit=built - 1)

    def test_a_search_range_replaces_the_network_mask(self):
        """With a wide mask, the search area is narrowed by a range."""
        candidates = ip_assign.search_candidates(
            "10.0.0.0", "255.0.0.0", first="10.1.1.10", last="10.1.1.12")
        self.assertEqual(candidates, ["10.1.1.10", "10.1.1.11", "10.1.1.12"])

    def test_a_broken_search_range_is_rejected(self):
        for first, last, expected in (
                ("10.1.1.10", "", "both the start and the end"),
                ("10.1.1.60", "10.1.1.10", "cannot be below its start"),
                ("10.1.0.0", "10.1.9.0", "too wide")):
            with self.assertRaises(ValueError,
                                   msg=f"{first}-{last}") as caught:
                ip_assign.search_candidates("", "", first=first, last=last)
            self.assertIn(expected, str(caught.exception))

    def test_the_run_tries_both_factory_addresses(self):
        """The fixed factory address plus the set-resolved one, both candidates.

        Before the factory address was fixed, the run searched 10.n.1.12; some
        field devices are still there. Looking only at the fixed address made
        them "not found".
        """
        from panel.ip_assign import runner

        topology = fakes.device_map([
            {"Name": "Intercom_1", "IP": "10.n.1.10", "IsActive": True,
             "Type": "Announcement", "SubType": "Intercom", "Port": "11",
             "Status": {}}], switch_ip="10.n.1.101")
        self.build_map(topology)
        inventory = device_map.load(8, self.map_path)
        switch = inventory.switches()[0]

        captured = {}

        def fake_execute(module, argv, emit, cancelled=None):
            captured["argv"] = argv
            return 0

        previous = runner._execute
        runner._execute = fake_execute
        self.addCleanup(lambda: setattr(runner, "_execute", previous))

        lines = []
        runner._run_intercom(inventory, switch, [11], ("admin", "x"),
                             lines.append, {})
        argv = captured["argv"]
        self.assertEqual(argv[argv.index("--factory-ip") + 1], "10.1.1.12")
        self.assertIn("--default-ip", argv)
        self.assertIn("10.8.1.12", argv[argv.index("--default-ip") + 1:])
        self.assertTrue(any("10.1.1.12" in line and "10.8.1.12" in line
                            for line in lines), lines)
        # The mask is not forced on the device: its own netmask is kept.
        self.assertNotIn("--force-netmask", argv)
        self.assertNotIn("--netmask", argv)

    def test_an_asked_for_mask_is_forced_past_the_devices_own(self):
        """The device normally keeps its netmask; an explicit choice wins.

        `write_ip` sends back whatever mask the device reports, so without
        --force-netmask the operator's answer would be quietly discarded.
        """
        from panel.ip_assign import runner

        inventory, switch, captured = self._intercom_run_capture()
        runner._run_intercom(inventory, switch, [11], ("admin", "x"),
                             lambda _line: None, {"targetPrefix": 8})
        argv = captured["argv"]
        self.assertEqual(argv[argv.index("--netmask") + 1], "255.0.0.0")
        self.assertIn("--force-netmask", argv)

    def test_a_project_that_states_a_mask_has_it_written(self):
        """Gaziray puts four cars on four /24s and the broker on a fifth.

        A device written with the system default /24 there can reach neither
        the broker at 10.n.0.1 nor the next car, so the project states /16 and
        the run must enforce it — the operator is not asked, and on the other
        projects nothing changes because they state nothing.
        """
        from panel.editions import runtime as editions
        from panel.ip_assign import runner

        inventory, switch, captured = self._intercom_run_capture()
        with mock.patch.object(editions, "prefix", return_value=16):
            runner._run_intercom(inventory, switch, [11], ("admin", "x"),
                                 lambda _line: None, {})
        argv = captured["argv"]
        self.assertEqual(argv[argv.index("--netmask") + 1], "255.255.0.0")
        self.assertIn("--force-netmask", argv)

    def test_the_operator_outranks_the_projects_mask(self):
        """A mask typed on the run form is the last word."""
        from panel.editions import runtime as editions
        from panel.ip_assign import runner

        inventory, switch, captured = self._intercom_run_capture()
        with mock.patch.object(editions, "prefix", return_value=16):
            runner._run_intercom(inventory, switch, [11], ("admin", "x"),
                                 lambda _line: None, {"targetPrefix": 8})
        argv = captured["argv"]
        self.assertEqual(argv[argv.index("--netmask") + 1], "255.0.0.0")

    def _intercom_run_capture(self):
        """One intercom, and `_execute` replaced by a recorder."""
        from panel.ip_assign import runner

        topology = fakes.device_map([
            {"Name": "Intercom_1", "IP": "10.n.1.10", "IsActive": True,
             "Type": "Announcement", "SubType": "Intercom", "Port": "11",
             "Status": {}}], switch_ip="10.n.1.101")
        self.build_map(topology)
        inventory = device_map.load(8, self.map_path)
        captured = {}

        def fake_execute(module, argv, emit, cancelled=None):
            captured["argv"] = argv
            return 0

        previous = runner._execute
        runner._execute = fake_execute
        self.addCleanup(lambda: setattr(runner, "_execute", previous))
        return inventory, inventory.switches()[0], captured

    def test_the_mask_is_read_in_both_spellings_and_bounded(self):
        self.assertEqual(ip_assign.parse_prefix(""), 24)
        self.assertEqual(ip_assign.parse_prefix("/8"), 8)
        self.assertEqual(ip_assign.parse_prefix("255.255.0.0"), 16)
        self.assertEqual(ip_assign.netmask_for(8), "255.0.0.0")
        for bad in ("0", "31", "255.0.255.0", "nonsense"):
            with self.assertRaises(ValueError, msg=bad):
                ip_assign.parse_prefix(bad)

    def test_a_mistyped_set_number_is_refused_rather_than_read_as_one(self):
        """The factory-reset scope: which set the devices are on NOW.

        Falling back to 1 would send the run looking on the wrong network and
        report every port as "device not found"."""
        self.assertEqual(ip_assign.parse_set(""), 0)
        self.assertEqual(ip_assign.parse_set("3"), 3)
        for bad in ("0", "255", "three", "-1"):
            with self.assertRaises(ValueError, msg=bad):
                ip_assign.parse_set(bad)

    def test_the_plan_comes_from_devicemap(self):
        topology = fakes.device_map([
            {"Name": "Intercom_1", "IP": "10.n.1.10", "IsActive": True,
             "Type": "Announcement", "SubType": "Intercom", "Port": "11",
             "Status": {}},
            {"Name": "Intercom_2", "IP": "10.n.1.11", "IsActive": True,
             "Type": "Announcement", "SubType": "Intercom", "Port": "12",
             "Status": {}},
        ], switch_ip="10.n.1.101")
        inventory = self.build_map(topology)
        plan = ip_assign.build_plan(inventory, "Intercom", [11, 12, 13])

        self.assertEqual(plan["targetCount"], 2)
        p11 = next(r for r in plan["rows"] if r["port"] == 11)
        self.assertEqual(p11["name"], "Intercom_1")
        self.assertEqual(p11["targetIp"], "10.1.1.10")
        p13 = next(r for r in plan["rows"] if r["port"] == 13)
        self.assertFalse(p13["actionable"])
        self.assertEqual(p13["name"], "—")

    def test_a_transfer_plan_shows_the_other_sets_addresses_as_the_source(self):
        topology = fakes.device_map([
            {"Name": "Intercom_1", "IP": "10.n.1.10", "IsActive": True,
             "Type": "Announcement", "SubType": "Intercom", "Port": "11",
             "Status": {}}], switch_ip="10.n.1.101")
        inventory = self.build_map(topology)

        plain = ip_assign.build_plan(inventory, "Intercom", [11])
        self.assertEqual(plain["rows"][0]["sourceIp"], "10.1.1.12")

        # The two sets are what the LCD factory reset turns around; the mask
        # is a plain run option.
        wide = ip_assign.build_plan(inventory, "Intercom", [11],
                                    source_set=3, target_prefix=8)
        self.assertEqual(wide["rows"][0]["sourceIp"], "10.3.1.10")
        self.assertEqual(wide["targetPrefix"], 8)
        self.assertEqual(wide["targetNetmask"], "255.0.0.0")

    def test_the_run_is_rejected_when_a_protected_port_is_targeted(self):
        """The computer's port or a switch link cannot enter the run.

        The run cycles PoE off and on; touching one of those ports cuts its own
        path and leaves it half done.
        """
        topology = fakes.device_map([
            {"Name": "Intercom_1", "IP": "10.n.1.10", "IsActive": True,
             "Type": "Announcement", "SubType": "Intercom", "Port": "11",
             "Status": {}}], switch_ip="10.n.1.101")
        inventory = self.build_map(topology)
        for reason in ("the computer is on this port",
                       "link to the other switch"):
            with self.assertRaises(ValueError) as caught:
                ip_assign.run(inventory, inventory.switches()[0].id, [11],
                              lambda line: None, protected={11: reason})
            self.assertIn("cut its own connection", str(caught.exception))
            self.assertIn(reason, str(caught.exception))


class ExcelExport(ServiceTest):

    def test_excel_writes_only_the_values_that_were_read(self):
        """The template is copied; an unread cell stays empty and the template
        is untouched."""
        inventory = device_map.load(1, YATAKLI_MAP)
        view = jobs.view_for(1)

        device = next(d for d in inventory.devices if d.read_method == "http")
        result = probe_result.success(
            {"version": "1.2.5", "serial": "SN-1", "uptime": "01:00:00"},
            "http")
        result.generation = jobs.next_generation()
        view.write(device.id, result)

        output = Path(tempfile.mkdtemp(prefix="excel-test-")) / "cikti.xlsx"
        template_before = settings.EXCEL_TEMPLATE.read_bytes()
        path = checklist.export(inventory, view.all(), output=output)

        self.assertTrue(path.is_file())
        self.assertGreater(path.stat().st_size, 5000)
        # The template file was not touched
        self.assertEqual(settings.EXCEL_TEMPLATE.read_bytes(),
                         template_before)

        import openpyxl
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[checklist.workbook.SHEET_NAME]
        verify = script_loader.device_verify()
        columns = {}
        for index in range(1, sheet.max_column + 1):
            column = cols.column_for_heading(
                sheet.cell(verify.HEADER_ROW, index).value)
            if column:
                columns[column] = index

        read_row = None
        for row in range(verify.HEADER_ROW + 1, sheet.max_row + 1):
            cell = sheet.cell(row, columns[verify.COL_IP_TEMPLATE]).value
            if cell == device.ip_template:
                read_row = row
                break
        self.assertIsNotNone(read_row)
        self.assertEqual(sheet.cell(read_row, columns[cols.VERSION]).value,
                         "1.2.5")
        self.assertEqual(
            sheet.cell(read_row, columns[cols.CONNECTION_INFO]).value, device.ip)

        # An unread device's version cell must stay EMPTY
        others = [d for d in inventory.devices
                  if d.id != device.id and view.get(d.id) is None]
        self.assertTrue(others)
        for row in range(verify.HEADER_ROW + 1, sheet.max_row + 1):
            template = sheet.cell(row, columns[verify.COL_IP_TEMPLATE]).value
            if template in {d.ip_template for d in others}:
                self.assertIsNone(
                    sheet.cell(row, columns[cols.VERSION]).value,
                    f"a made-up version was written for {template}")


class EveryShippedDeviceMapIsUnderstood(unittest.TestCase):
    """A device kind in a delivered map that no table knows is a silent hole.

    The fuar (exhibition) map arrived with three kinds this panel had never
    been shown — a Swanneck microphone, a Twin LCD, and cameras called Front
    and Cabin — and every one of them would have loaded, drawn a row in the
    device list, and then been reachable from nowhere. Nothing would have
    said so: `read_method_for` answers "mqtt" for anything it does not
    recognise, which means "described by DeviceMap, asked nothing", and that
    is indistinguishable on screen from a device that is genuinely passive.
    """

    def maps(self):
        found = sorted((settings.ROOT / "devicemaps").rglob("DeviceMap*.json"))
        here = [path for path in found if path.is_file()]
        # Named, so a checkout that has lost the maps fails loudly instead of
        # passing both checks below with nothing to check.
        self.assertTrue(here, "no DeviceMap in this checkout")
        return here

    def test_every_device_is_reached_or_deliberately_passive(self):
        # The kinds that really are DeviceMap-only, each with its reason.
        # NOT a place to park a device nobody has wired up yet.
        passive = {
            ("LED", "Front"): "a display driven by the PISCU, not addressed",
            ("LED", "Side"): "the side destination display, driven by the "
                             "PISCU like the front one",
            ("LCD", "Landing"): "a passive display; nothing to ask it",
            ("ICU", None): "reports through the PISCU",
            ("AP", None): "the access point is commissioned by hand",
            ("Router", "Twin"): "commissioned by hand, like the access "
                                "point; the panel neither reads nor writes it",
            ("Announcement", "Induction Loop"):
                "a hearing-loop amplifier with no API to ask — it carries a "
                "PBX extension but answers nothing the panel can query",
        }
        unreached = []
        for path in self.maps():
            data = json.loads(path.read_text(encoding="utf-8"))
            for switch in data.get("Switches") or []:
                for device in switch.get("Devices") or []:
                    pair = (device.get("Type"), device.get("SubType") or None)
                    if catalog.read_method_for(*pair) != "mqtt":
                        continue
                    if pair in passive:
                        continue
                    unreached.append(f"{path.name}: {pair[0]}/{pair[1] or '-'}")
        self.assertEqual(sorted(set(unreached)), [],
                         "a device kind no table knows — give it a read "
                         "method in panel/inventory/catalog.py, or a line in "
                         "`passive` above saying why it has none")

    def test_every_reachable_device_belongs_to_a_group(self):
        """A read method alone reaches nothing: the screens work on GROUPS,
        and a device in no group cannot be selected on any of them."""
        orphans = []
        for path in self.maps():
            inventory = device_map.load(1, path, cache=False)
            for device in inventory.devices:
                if device.type == "Switch" or device.read_method == "mqtt":
                    continue
                if any(catalog.device_supports(device, op)
                       for op in ("ip", "cfg", "fw", "check")):
                    continue
                orphans.append(f"{path.name}: {device.type}/"
                               f"{device.subtype or '-'}")
        self.assertEqual(sorted(set(orphans)), [],
                         "reachable, but in no group — add one to "
                         "catalog.GROUPS")

    # What each project's target pickers may offer. Written out rather than
    # recomputed: the whole point is that these are the customer's own device
    # types, and a table that derived itself from the same function it checks
    # would keep passing while the lists silently grew back.
    #
    # These were the numbers before `groups_for` existed: every project was
    # handed all eleven `cfg` groups and all nine `fw` ones, so a Yatakli
    # operator's picker listed the exhibition rack's Twin LCD and GDM's two
    # screens, and VIP's listed a Compartment LCD it does not have.
    OFFERED: ClassVar[dict] = {
        "DeviceMap_Yatakli.json": {
            "cfg": ["Intercom", "Handset", "Amplifier", "UIC",
                    "Compartment LCD", "Camera", "NVR"],
            "fw": ["Intercom", "Handset", "Amplifier", "UIC",
                   "Compartment LCD"],
            "ip": ["Intercom", "Compartment LCD"],
        },
        "DeviceMap_Vip.json": {
            "cfg": ["Intercom", "Handset", "Amplifier", "UIC", "Camera",
                    "NVR"],
            "fw": ["Intercom", "Handset", "Amplifier", "UIC"],
            "ip": ["Intercom"],
        },
        "DeviceMap_Gdm.json": {
            "cfg": ["Intercom", "Amplifier", "Swanneck", "LINE LCD",
                    "PIS LCD", "Camera", "NVR"],
            "fw": ["Intercom", "Amplifier", "Swanneck", "LINE LCD",
                   "PIS LCD"],
            "ip": ["Intercom"],
        },
        "DeviceMap_Gaziray.json": {
            "cfg": ["Intercom", "Amplifier", "Swanneck", "UIC", "Twin LCD",
                    "Camera", "NVR"],
            "fw": ["Intercom", "Amplifier", "Swanneck", "UIC", "Twin LCD"],
            "ip": ["Intercom"],
        },
        "DeviceMap_Fuar.json": {
            "cfg": ["Intercom", "Handset", "Amplifier", "Swanneck",
                    "Twin LCD", "Camera"],
            "fw": ["Intercom", "Handset", "Amplifier", "Swanneck",
                   "Twin LCD"],
            "ip": ["Intercom"],
        },
    }

    def test_a_picker_offers_only_this_project_s_own_device_types(self):
        """`groups_for` answers with the customer's equipment and no one
        else's.

        The screens take this list verbatim (`/api/project` → `meta.groups` →
        `group_bar.groupsFor`), so what is offered here is what an operator
        can point an operation at.
        """
        for path in self.maps():
            inventory = device_map.load(1, path, cache=False)
            offered = catalog.groups_for(inventory)
            with self.subTest(project=path.name):
                # Nothing is offered that the project has no device for.
                empty = [group["name"] for group in offered
                         if not any(catalog.group_matches(group, device)
                                    for device in inventory.devices)]
                self.assertEqual(empty, [], "offered with nothing behind it")
                expected = self.OFFERED.get(path.name)
                self.assertIsNotNone(
                    expected, f"{path.name} is new — add its row to OFFERED")
                for op, names in expected.items():
                    self.assertEqual(
                        [group["name"] for group in offered
                         if catalog.group_supports(group, op)],
                        names, op)

    def test_a_group_this_project_does_not_have_is_refused(self):
        """The picker no longer offering it is not the same as the API
        refusing it: the client may still hold the list from the project that
        was open a moment ago, and these are the write paths."""
        vip = next(path for path in self.maps()
                   if path.name == "DeviceMap_Vip.json")
        inventory = device_map.load(1, vip, cache=False)
        # Yatakli has these; VIP does not.
        for absent in ("Compartment LCD", "Twin LCD"):
            with self.subTest(group=absent):
                self.assertIsNotNone(catalog.find_group(absent),
                                     "the vocabulary still knows it")
                self.assertIsNone(catalog.group_in(inventory, absent))
                self.assertEqual(ip_assign.resolve_groups([absent], inventory),
                                 [])
        # And what VIP does have still resolves.
        self.assertIsNotNone(catalog.group_in(inventory, "Intercom"))


class ReadMethodsMatchTheFieldScript(unittest.TestCase):
    """`read_method_for` and device_verify's COLLECTORS must not drift apart.

    Two tables describe how a device is reached. `panel.inventory.catalog`
    names the method the panel reads and writes with; the COLLECTORS table
    in `field_scripts/device_verify.py` names the function the checklist
    export queries with. They answer different questions and are not merged
    — but a device that one of them thinks is reachable and the other has
    never heard of is a silent hole, and the first sign of it is an empty
    Excel column nobody can explain.

    The claim that the two mirror each other used to sit in a docstring,
    where nothing could check it. It was already false: Announcement/UIC.
    """

    # The panel names a method; the field script names the function that
    # implements it. This is the join between the two tables.
    METHOD_FOR_COLLECTOR: ClassVar[dict] = {
        "fetch_switch": "kyland",
        "fetch_isapi": "isapi",
        "fetch_announcement": "http",
        "fetch_piscu": "app",
        "fetch_hmi": "app",
        "fetch_compartment_lcd": "adb",
    }

    # A device the panel queries but the field script deliberately does not.
    # Same rule as the allowlists in tests/test_language.py: narrow, and
    # every entry says why. NOT somewhere to park a collector nobody has
    # written yet.
    NO_COLLECTOR_ON_PURPOSE: ClassVar[dict] = {
        ("Announcement", "UIC"): (
            "the panel writes its SIP and threshold settings over HTTP "
            "(panel/config_sync/targets.py), but every field the checklist "
            "reports for a UIC already arrives in DeviceMap, so the export "
            "has no second query to make"),
    }

    # `mqtt` is the fallback in `read_method_for`: it means the device is
    # described by DeviceMap and asked nothing directly.
    FED_FROM_DEVICEMAP = "mqtt"

    def setUp(self):
        script = script_loader.device_verify()
        self.collectors = script.COLLECTORS
        # The same lookup the export itself makes — a key with None for the
        # subtype answers for every subtype of that type (see the table's own
        # note). Asking the dict directly here would fail a project whose
        # cameras are called Front and Cabin while the export reads them
        # perfectly well.
        self.collector_for = script.collector_for

    def device_pairs(self):
        """Every (type, subtype) in the DeviceMaps this checkout carries.

        Read from the files rather than a fixture: the point is to catch a
        device the customer really has and the tables really disagree on.
        """
        pairs = set()
        maps = sorted((settings.ROOT / "devicemaps").rglob("DeviceMap*.json"))
        found = 0
        for path in maps:
            if not path.is_file():
                continue
            found += 1
            data = json.loads(path.read_text(encoding="utf-8"))
            for switch in data.get("Switches") or []:
                pairs.add(("Switch", None))
                for device in switch.get("Devices") or []:
                    pairs.add((device.get("Type"),
                               device.get("SubType") or None))
        self.assertTrue(found, "no DeviceMap to check the tables against")
        return pairs

    def test_every_collector_agrees_with_the_read_method_map(self):
        for (kind, subtype), collector in sorted(
                self.collectors.items(), key=str):
            expected = self.METHOD_FOR_COLLECTOR.get(collector.__name__)
            self.assertIsNotNone(
                expected,
                f"{collector.__name__} is a new collector; say which read "
                "method it implements in METHOD_FOR_COLLECTOR")
            self.assertEqual(catalog.read_method_for(kind, subtype), expected,
                             f"{kind}/{subtype}")

    def test_every_queried_device_has_a_collector_or_a_stated_reason(self):
        missing = []
        for kind, subtype in sorted(self.device_pairs(), key=str):
            method = catalog.read_method_for(kind, subtype)
            if method == self.FED_FROM_DEVICEMAP:
                continue
            if self.collector_for(kind, subtype):
                continue
            if (kind, subtype) in self.NO_COLLECTOR_ON_PURPOSE:
                continue
            missing.append(f"{kind}/{subtype or '-'} reads over {method!r} "
                           "but device_verify has no collector for it")
        self.assertEqual(
            missing, [],
            "the two tables disagree. Add a collector to "
            "field_scripts/device_verify.py, or an entry saying why not to "
            "NO_COLLECTOR_ON_PURPOSE:\n  " + "\n  ".join(missing))

    def test_the_exception_list_is_still_needed(self):
        """A stale exemption hides the next disagreement."""
        stale = []
        for kind, subtype in self.NO_COLLECTOR_ON_PURPOSE:
            if self.collector_for(kind, subtype):
                stale.append(f"{kind}/{subtype}: device_verify collects it "
                             "now, so the exemption is wrong")
            elif catalog.read_method_for(kind, subtype) == (
                    self.FED_FROM_DEVICEMAP):
                stale.append(f"{kind}/{subtype}: the panel no longer queries "
                             "it, so no exemption is needed")
        self.assertEqual(stale, [], "stale exemptions:\n  " + "\n  ".join(stale))


if __name__ == "__main__":
    unittest.main()
