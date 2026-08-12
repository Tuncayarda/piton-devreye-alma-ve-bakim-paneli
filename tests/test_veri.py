#!/usr/bin/env python3
"""Veri sunumu, envanter ve Excel çıktısı.

Buradaki testlerin ortak derdi tek: panel hiçbir yerde olmayan bir veriyi
varmış gibi göstermemeli. "Okunmadı" ile "uygulanmıyor" ayrı, sahte
varsayılan yok, Excel'e yalnız gerçekten okunan değer yazılıyor.
"""
from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from core import device_map, dogrulama, excel, ip_atama, isler, kategori, okuma

from .ortak import PanelTesti, ServisTesti
from . import sahte


class Envanter(PanelTesti):

    def test_ip_sablonundaki_n_cozulur(self):
        self.assertEqual(device_map.coz("10.n.1.24", 3), "10.3.1.24")
        self.assertEqual(device_map.coz("10.n.1.24", 12), "10.12.1.24")
        # 'n' harfi başka bir kelimenin içindeyse dokunulmaz
        self.assertEqual(device_map.coz("10.no.1.24", 3), "10.no.1.24")
        self.assertEqual(device_map.sablonla("10.3.1.24", 3), "10.n.1.24")

    def test_gercek_devicemap_butun_setlerde_cozulur(self):
        from core import ayar
        for n in (1, 2, 9, 16):
            env = device_map.yukle(n, ayar.KOK / "DeviceMap.json")
            self.assertEqual(env.set_no, n)
            self.assertGreater(len(env.cihazlar), 30)
            for c in env.cihazlar:
                self.assertNotIn(".n.", c.ip, f"{c.ad} şablonu çözülmemiş")
                self.assertTrue(c.ip.startswith(f"10.{n}."), c.ip)

    def test_dto_kimlik_alanlari_icermez(self):
        from core import ayar
        env = device_map.yukle(1, ayar.KOK / "DeviceMap.json")
        for c in env.cihazlar:
            veri = c.dto()
            self.assertNotIn("Password", str(veri))
            self.assertNotIn("Username", str(veri))
            self.assertNotIn("password", str(veri).lower())
            for alan in ("id", "ad", "ip", "ipSablonu", "type", "subtype",
                         "switch", "port", "kategori", "yontem"):
                self.assertIn(alan, veri, alan)
            self.assertNotIn("Password", c.ekstra)
            self.assertNotIn("Username", c.ekstra)

    def test_cihaz_id_leri_benzersiz_ve_kararli(self):
        from core import ayar
        a = device_map.yukle(1, ayar.KOK / "DeviceMap.json")
        b = device_map.yukle(4, ayar.KOK / "DeviceMap.json")
        idler = [c.id for c in a.cihazlar]
        self.assertEqual(len(idler), len(set(idler)))
        # id set numarasından bağımsız, IP değişir
        self.assertEqual(idler, [c.id for c in b.cihazlar])
        self.assertNotEqual([c.ip for c in a.cihazlar],
                            [c.ip for c in b.cihazlar])


class VeriSunumu(PanelTesti):

    def test_okunmadi_ile_uygulanmiyor_ayridir(self):
        okunmadi = dogrulama.okunmadi("mqtt")
        uygulanmaz = dogrulama.uygulanmaz("mqtt", "Bu cihazda uygulanmıyor")
        self.assertEqual(okunmadi.durum, dogrulama.GRI)
        self.assertEqual(uygulanmaz.durum, dogrulama.GRI)
        self.assertNotEqual(okunmadi.dogrulama, uygulanmaz.dogrulama)
        self.assertEqual(okunmadi.dogrulama, dogrulama.OKUNMADI)
        self.assertEqual(uygulanmaz.dogrulama, dogrulama.UYGULANMIYOR)

    def test_okunmayan_alanlarda_sahte_varsayilan_yok(self):
        """Cihaz cevap vermezse hiçbir alan uydurulmaz."""
        harita = sahte.device_map([], switch_ip="127.0.0.1")
        env = self.kur_harita(harita)
        with sahte.sessiz() as s:
            self.switch_portu(s.port)
            sonuc = okuma.cihaz_oku(env.switchler()[0], kimlik=("a", "b"),
                                    timeout=1.0)
        self.assertEqual(sonuc.durum, dogrulama.KIRMIZI)
        self.assertEqual(sonuc.alanlar, {})

    def test_mqtt_yoksa_cihaz_gri_kalir_hatali_gorunmez(self):
        """Broker yoksa MQTT kaynaklı cihaz 'okunmadı' olur, 'hatalı' değil."""
        harita = sahte.device_map([{
            "Name": "Ap_1", "IP": "127.0.0.1", "IsActive": True,
            "Type": "AP", "SubType": "", "Port": "26", "Status": {},
        }])
        env = self.kur_harita(harita)
        ap = env.tip_ile("AP")[0]
        sonuc = okuma.cihaz_oku(ap, telemetri=None)
        self.assertEqual(sonuc.durum, dogrulama.GRI)
        self.assertEqual(sonuc.dogrulama, dogrulama.OKUNMADI)
        self.assertIn("MQTT", sonuc.aciklama)

    def test_sayaclar_renk_basina_ayri(self):
        say = dogrulama.sayilar([
            dogrulama.basarili({}, "http"),
            dogrulama.basarili({}, "http"),
            dogrulama.Sonuc(durum=dogrulama.TURUNCU),
            dogrulama.Sonuc(durum=dogrulama.KIRMIZI),
            dogrulama.okunmadi("mqtt"),
        ])
        self.assertEqual(say, {"basarili": 2, "erisimBekleyen": 1,
                               "hatali": 1, "okunmayan": 1})

    def test_uptime_metni(self):
        self.assertEqual(dogrulama.uptime_metni(3661), "01:01:01")
        self.assertEqual(dogrulama.uptime_metni(-1), "")
        self.assertEqual(dogrulama.uptime_metni(None), "")
        self.assertEqual(dogrulama.uptime_metni("abc"), "")


class AnonsAlanlari(PanelTesti):
    """Gain ve arama numarası ayrı alanlardır, ses seviyesi değildir."""

    def kur(self):
        harita = sahte.device_map([{
            "Name": "Intercom_1", "IP": "127.0.0.1", "IsActive": True,
            "Type": "Announcement", "SubType": "Intercom", "Port": "11",
            "PBXExtension": "2001", "Status": {"NoError": True},
        }])
        return self.kur_harita(harita).tip_ile("Announcement")[0]

    def test_gain_ve_arama_no_okunur(self):
        from core import ayar
        cihaz = self.kur()
        with sahte.anons() as an:
            ayar.ANONS_PORT = an.port
            sonuc = okuma.cihaz_oku(cihaz)
        a = sonuc.alanlar
        self.assertEqual(sonuc.durum, dogrulama.YESIL)
        self.assertEqual(a["hoparlorGain"], 4)
        self.assertEqual(a["mikrofonGain"], 2)
        self.assertEqual(a["sipArama"], "5001")
        # Gain, ses seviyesinin yerine geçmez
        self.assertEqual(a["hoparlor"], 70)
        self.assertEqual(a["mikrofon"], 60)
        # Arama numarası cihazın kendi dahilisiyle karışmaz
        self.assertEqual(a["sipDahili"], "2001")

    def test_alan_yoksa_uydurulmaz(self):
        from core import ayar
        cihaz = self.kur()
        eksik = {"firmwareVersion": "1.2.5", "uptime": 10,
                 "pbxExtension": "2001"}
        with sahte.anons(ayarlar=eksik) as an:
            ayar.ANONS_PORT = an.port
            a = okuma.cihaz_oku(cihaz).alanlar
        self.assertIsNone(a["hoparlorGain"])
        self.assertIsNone(a["mikrofonGain"])
        self.assertEqual(a["sipArama"], "")

    def test_excel_eslemesi_yeni_sutunlari_kapsar(self):
        for alan, baslik in (("hoparlorGain", "Hoparlör Gain"),
                             ("mikrofonGain", "Mikrofon Gain"),
                             ("sipArama", "SIP Arama No")):
            self.assertEqual(excel.ESLEME.get(alan), baslik, alan)


class TelemetriEslesmesi(PanelTesti):
    """Broker şablon IP yayınlar, panel çözülmüş IP ile arar.

    İkisi eşleşmediğinde MQTT kaynaklı bütün cihazlar "telemetride
    bulunamadı" diye kırmızıya düşüyordu.
    """

    def kur_telemetri(self, set_no=1):
        from core import piscu
        t = piscu.Telemetri("10.1.1.1")
        t.harita.clear()
        t._ekle("10.n.1.4", {"Name": "Hmi", "IP": "10.n.1.4",
                             "Status": {"NoError": True, "Uptime": 27241}},
                set_no)
        t.uygulama = {"ClientManager_MCP_YATAKLI_1": {
            "ClientId": "ClientManager_MCP_YATAKLI_1", "DeviceIP": "10.1.1.4",
            "HWID": "34DA8534", "Status": "connected", "Version": "1.2.5"}}
        return t

    def test_sablon_ip_cozulmus_ip_ile_bulunur(self):
        t = self.kur_telemetri()
        self.assertIsNotNone(t.kayit("10.1.1.4"))      # çözülmüş
        self.assertIsNotNone(t.kayit("10.n.1.4"))      # şablon
        self.assertIsNone(t.kayit("10.1.1.9"))

    def test_hmi_calisma_suresi_devicemapten_tamamlanir(self):
        """AppStatus yükünde Uptime yok; süre DeviceMap kaydından gelir."""
        harita = sahte.device_map([{
            "Name": "Hmi", "IP": "10.n.1.4", "IsActive": True,
            "Type": "HMI", "SubType": "", "Port": "6", "Status": {},
        }], switch_ip="10.n.1.100")
        env = self.kur_harita(harita)
        hmi = env.tip_ile("HMI")[0]
        sonuc = okuma.cihaz_oku(hmi, telemetri=self.kur_telemetri())
        self.assertEqual(sonuc.durum, dogrulama.YESIL)
        self.assertEqual(sonuc.alanlar["surum"], "1.2.5")
        self.assertEqual(sonuc.alanlar["calisma"], "07:34:01")

    def test_uptime_eksi_bir_calisma_suresi_sayilmaz(self):
        """PISCU kapalı cihaz için -1 bildiriyor; bu bir süre değildir."""
        harita = sahte.device_map([{
            "Name": "Hmi", "IP": "10.n.1.4", "IsActive": True,
            "Type": "HMI", "SubType": "", "Port": "6", "Status": {},
        }], switch_ip="10.n.1.100")
        env = self.kur_harita(harita)
        t = self.kur_telemetri()
        t.harita["10.1.1.4"]["Status"]["Uptime"] = -1
        sonuc = okuma.cihaz_oku(env.tip_ile("HMI")[0], telemetri=t)
        self.assertEqual(sonuc.alanlar["calisma"], "")


class RetainedMesajCihazVarligiDegildir(PanelTesti):
    """Broker'daki retained mesaj cihaz gittikten sonra da orada durur.

    Sahada görülen: HMI kablosu takılı değilken panel onu yeşil
    "Doğrulandı" gösteriyordu. Cihazın IP'si hiç cevap vermiyordu; panel
    ise broker'da kalan mesajı okuyup ayakta sanıyordu. Üstelik gösterdiği
    notun kendisi "disconnected" yazıyordu.

    Buradaki yükler sahadaki broker'dan birebir alınmıştır.
    """

    HMI = {
        "Name": "Hmi", "IP": "10.n.1.4", "IsActive": True,
        "Type": "HMI", "SubType": "", "Port": "6", "Status": {},
    }
    ICU = {
        "Name": "Icu", "IP": "10.n.1.2", "IsActive": True,
        "Type": "ICU", "SubType": "", "Port": "5", "Status": {},
    }

    def _oku(self, canli, uygulama, tanim=None):
        """Verilen broker görüntüsüyle tek cihazı okur.

        `canli`   ALFA/DeviceMap kaydı (None: kayıt hiç yok)
        `uygulama` ALFA/AppStatus mesajları {ClientId: yük}
        """
        from core import piscu

        tanim = tanim or self.HMI
        env = self.kur_harita(sahte.device_map([tanim],
                                               switch_ip="10.n.1.100"))
        t = piscu.Telemetri("10.1.1.1")
        t.harita.clear()
        if canli is not None:
            t._ekle(tanim["IP"], canli, 1)
        t.uygulama = uygulama
        return okuma.cihaz_oku(env.tip_ile(tanim["Type"])[0], telemetri=t)

    # ── AppStatus: son vasiyet (LWT) ──
    def test_disconnected_lwt_yesile_gecmez(self):
        """Sahadaki yük: {"ClientId": "...MCP...", "Status": "disconnected"}

        DeviceIP, HWID, Version alanlarının hiçbiri yok. Panel bu kaydı
        ClientId'ye bakarak buluyor ve boş alanlarla "doğrulandı" diyordu.
        """
        sonuc = self._oku(
            {"IP": "10.n.1.4", "Status": {"NoError": False, "Uptime": -1}},
            {"ClientManager_MCP_YATAKLI_1": {
                "ClientId": "ClientManager_MCP_YATAKLI_1",
                "Status": "disconnected"}})
        self.assertEqual(sonuc.durum, dogrulama.KIRMIZI)
        self.assertIn("bağlı değil", sonuc.aciklama)
        self.assertIn("disconnected", sonuc.aciklama)

    def test_bagli_cihaz_yesil_kalir(self):
        """Düzeltme çalışan cihazı bozmamalı."""
        sonuc = self._oku(
            {"IP": "10.n.1.4", "Status": {"NoError": True, "Uptime": 27241}},
            {"ClientManager_MCP_YATAKLI_1": {
                "ClientId": "ClientManager_MCP_YATAKLI_1",
                "DeviceIP": "10.1.1.4", "HWID": "34DA8534",
                "Status": "connected", "Version": "1.2.5"}})
        self.assertEqual(sonuc.durum, dogrulama.YESIL)
        self.assertEqual(sonuc.alanlar["surum"], "1.2.5")

    def test_appstatus_connected_dese_de_canli_kayit_arizali_ise_yesil_olmaz(self):
        """İki işaret çelişirse ayakta sayılmaz.

        AppStatus mesajı cihaz gittikten sonra "connected" kalmış olabilir
        (vasiyet yayınlanamadan güç kesilirse). PISCU'nun canlı kaydı ise
        cihazı o an izliyor.
        """
        sonuc = self._oku(
            {"IP": "10.n.1.4", "Status": {"NoError": False, "Uptime": -1}},
            {"ClientManager_MCP_YATAKLI_1": {
                "ClientId": "ClientManager_MCP_YATAKLI_1",
                "DeviceIP": "10.1.1.4", "HWID": "34DA8534",
                "Status": "connected", "Version": "1.2.5"}})
        self.assertEqual(sonuc.durum, dogrulama.KIRMIZI)
        self.assertEqual(
            sonuc.aciklama,
            "PISCU cihazın kapalı veya arızalı olduğunu bildiriyor")

    def test_status_yazmayan_eski_yuk_canli_kayitla_degerlendirilir(self):
        """AppStatus'ta Status alanı yoksa tek işaret canlı kayıttır."""
        sonuc = self._oku(
            {"IP": "10.n.1.4", "Status": {"NoError": True, "Uptime": 100}},
            {"ClientManager_MCP_YATAKLI_1": {
                "ClientId": "ClientManager_MCP_YATAKLI_1",
                "DeviceIP": "10.1.1.4", "Version": "1.2.5"}})
        self.assertEqual(sonuc.durum, dogrulama.YESIL)

    # ── canlı DeviceMap kaynaklı cihazlar (ICU, AP, LED, Landing LCD) ──
    def test_arizali_cihaz_uygulanmiyor_degil_hata_olur(self):
        """Kapalı cihaz "N/A" (gri) gösteriliyordu.

        Gri "bu cihazda bu denetim yok" demek. Burada denetim yapıldı ve
        cihaz arızalı bildirildi — bu bir sonuçtur, yokluk değil.
        """
        sonuc = self._oku(
            {"IP": "10.n.1.2", "Status": {"NoError": False, "Uptime": -1}},
            {}, tanim=self.ICU)
        self.assertEqual(sonuc.durum, dogrulama.KIRMIZI)
        self.assertEqual(sonuc.dogrulama, dogrulama.DOGRULANAMADI)
        self.assertEqual(
            sonuc.aciklama,
            "PISCU cihazın kapalı veya arızalı olduğunu bildiriyor")

    def test_saglam_mqtt_cihazi_yesil_kalir(self):
        sonuc = self._oku(
            {"IP": "10.n.1.2", "SerialNumber": "ICU-9",
             "Status": {"NoError": True, "Uptime": 500, "Version": "2.0"}},
            {}, tanim=self.ICU)
        self.assertEqual(sonuc.durum, dogrulama.YESIL)
        self.assertEqual(sonuc.alanlar["seri"], "ICU-9")


# ── Compartment LCD / ADB ───────────────────────────────────────────────
DUMPSYS = """Packages:
  Package [com.piton.train_lcd_panel] (5b3a8c0):
    userId=10123
    codePath=/data/app/com.piton.train_lcd_panel-1
    versionCode=1 minSdk=21 targetSdk=35
    versionName=0.0.5
    flags=[ HAS_CODE ALLOW_CLEAR_USER_DATA ]
    firstInstallTime=2026-07-07 13:13:40
    lastUpdateTime=2026-07-28 08:05:32
"""

LOGCAT = """08-05 12:04:17.249 13854 13946 I AnnounceSip: Registration state=registered code=200
08-05 12:24:25.747 13465 13557 I AnnounceSip: SIP engine started: sip:6001@10.1.1.1:5060 (UDP)
08-05 12:24:25.757 13465 13557 I AnnounceSip: Registration state=registered code=200
"""


class SahteAdb:
    """subprocess.run yerine geçer; adb çağrılarına hazır çıktı döndürür."""

    def __init__(self, **cevaplar):
        self.cevaplar = cevaplar
        self.cagrilar: list[list[str]] = []

    def __call__(self, args, **kwargs):
        self.cagrilar.append(list(args))
        birlesik = " ".join(args)
        cikti = ""
        if "connect" in args:
            cikti = "connected"
        elif "ro.serialno" in birlesik:
            cikti = self.cevaplar.get("seri", "rk3568r0001")
        elif "persist.sys.timezone" in birlesik:
            cikti = self.cevaplar.get("tz", "CST-3:00:00")
        elif "/proc/uptime" in birlesik:
            cikti = self.cevaplar.get("uptime", "3661.20 12000.00")
        elif "dumpsys" in args:
            cikti = self.cevaplar.get("dumpsys", DUMPSYS)
        elif "logcat" in args:
            cikti = self.cevaplar.get("logcat", LOGCAT)

        class Sonuc:
            stdout = cikti
            stderr = ""
            returncode = 0
        return Sonuc()


class CompartmentLcd(PanelTesti):
    """ADB okuması: bağlanmak yetmez, veri de gelmeli.

    Cihaza adb ile bağlanıp uygulama sürümünü okuyamadığımız halde yeşil
    göstermek, sahada kurulmamış uygulamayı kurulmuş gibi raporluyordu.
    """

    def kur_lcd(self):
        harita = sahte.device_map([{
            "Name": "Compartment_Lcd_1", "IP": "10.n.1.40", "IsActive": True,
            "Type": "LCD", "SubType": "Compartment", "Port": "13",
            "PBXExtension": "6001", "Status": {},
        }])
        env = self.kur_harita(harita)
        return env.tip_ile("LCD")[0]

    def yamala(self, sahte_adb):
        eski = okuma.subprocess.run
        okuma.subprocess.run = sahte_adb
        self.addCleanup(lambda: setattr(okuma.subprocess, "run", eski))

    def test_dumpsys_surumu_ayristirilir(self):
        p = okuma.paket_bilgisi(DUMPSYS)
        self.assertEqual(p["surum"], "0.0.5")
        self.assertEqual(p["surumKodu"], "1")
        self.assertEqual(p["minSdk"], "21")
        self.assertEqual(p["hedefSdk"], "35")
        self.assertEqual(p["kurulum"], "2026-07-07 13:13:40")
        self.assertEqual(p["guncelleme"], "2026-07-28 08:05:32")

    def test_paket_kurulu_degilse_surum_bos(self):
        p = okuma.paket_bilgisi("Unable to find package: com.piton.train_lcd_panel")
        self.assertEqual(p["surum"], "")

    def test_sip_gunlugunden_dahili_ve_kayit(self):
        s = okuma.sip_gunlugu(LOGCAT)
        self.assertEqual(s["sipDahili"], "6001")
        self.assertEqual(s["sipPbx"], "10.1.1.1")
        self.assertEqual(s["sipPort"], "5060")
        self.assertEqual(s["sipTasima"], "UDP")
        self.assertEqual(s["sipKayit"], "registered")
        self.assertEqual(s["sipKod"], "200")

    def test_sip_satiri_yoksa_alanlar_bos_kalir(self):
        s = okuma.sip_gunlugu("")
        self.assertEqual(s["sipDahili"], "")
        self.assertEqual(s["sipKayit"], "")

    def test_okunan_lcd_yesil_ve_alanlari_dolu(self):
        cihaz = self.kur_lcd()
        self.yamala(SahteAdb())
        sonuc = okuma.cihaz_oku(cihaz)
        self.assertEqual(sonuc.durum, dogrulama.YESIL)
        self.assertEqual(sonuc.alanlar["surum"], "0.0.5")
        self.assertEqual(sonuc.alanlar["seri"], "rk3568r0001")
        self.assertEqual(sonuc.alanlar["sipDahili"], "6001")
        self.assertEqual(sonuc.alanlar["sipPbx"], "10.1.1.1")
        self.assertEqual(sonuc.alanlar["sipKayit"], "registered (200)")
        self.assertEqual(sonuc.alanlar["calisma"], "01:01:01")
        self.assertEqual(sonuc.alanlar["saatDilimi"], "CST-3:00:00")

    def test_surum_okunamayan_lcd_yesil_gorunmez(self):
        """Asıl şikâyet: bilgi alınamadığı halde 'Aktif' görünüyordu."""
        cihaz = self.kur_lcd()
        self.yamala(SahteAdb(dumpsys="Unable to find package: com.piton.x"))
        sonuc = okuma.cihaz_oku(cihaz)
        self.assertEqual(sonuc.durum, dogrulama.KIRMIZI)
        self.assertEqual(sonuc.dogrulama, dogrulama.DOGRULANAMADI)
        self.assertEqual(sonuc.alanlar, {})
        self.assertIn("sürümü okunamadı", sonuc.aciklama)

    def test_baglanamayan_lcd_kirmizi(self):
        cihaz = self.kur_lcd()
        self.yamala(SahteAdb(seri=""))
        sonuc = okuma.cihaz_oku(cihaz)
        self.assertEqual(sonuc.durum, dogrulama.KIRMIZI)
        self.assertIn("adb", sonuc.aciklama)

    def test_sip_kaydi_yoksa_okuma_yine_de_gecerli(self):
        """Günlük tamponu dönmüş olabilir; SIP satırının yokluğu hata değil."""
        cihaz = self.kur_lcd()
        self.yamala(SahteAdb(logcat=""))
        sonuc = okuma.cihaz_oku(cihaz)
        self.assertEqual(sonuc.durum, dogrulama.YESIL)
        self.assertEqual(sonuc.alanlar["surum"], "0.0.5")
        self.assertEqual(sonuc.alanlar["sipDahili"], "")

    def test_gunlukte_yoksa_dahili_broker_duyurusundan_gelir(self):
        """Uygulamayı yeniden başlatmadan dahili numara okunabilmeli."""
        from core import piscu
        cihaz = self.kur_lcd()
        t = piscu.Telemetri("10.1.1.1")
        t.sip = {cihaz.ip: "6001"}

        self.yamala(SahteAdb(logcat=""))
        sonuc = okuma.cihaz_oku(cihaz, telemetri=t)
        self.assertEqual(sonuc.alanlar["sipDahili"], "6001")
        self.assertIn("ALFA/SipPort", sonuc.alanlar["sipDahiliKaynak"])

        # Günlükte varsa cihazın kendi beyanı tercih edilir
        self.yamala(SahteAdb())
        sonuc = okuma.cihaz_oku(cihaz, telemetri=t)
        self.assertEqual(sonuc.alanlar["sipDahili"], "6001")
        self.assertEqual(sonuc.alanlar["sipDahiliKaynak"], "cihaz günlüğü")

    def test_pbx_adresi_kayitliyken_projeden_tamamlanir(self):
        """Günlükteki `sip:...@10.1.1.1` satırı düşmüşse PBX setin PISCU'su."""
        cihaz = self.kur_lcd()
        self.yamala(SahteAdb(logcat=(
            "08-05 12:04:17.249 I AnnounceSip: "
            "Registration state=registered code=200\n")))
        a = okuma.cihaz_oku(cihaz, pbx_ip="10.1.1.1").alanlar
        self.assertEqual(a["sipPbx"], "10.1.1.1")
        self.assertEqual(a["sipPbxKaynak"], "proje (PISCU)")

        # Günlükte varsa cihazın kendi beyanı tercih edilir
        self.yamala(SahteAdb())
        a = okuma.cihaz_oku(cihaz, pbx_ip="10.9.9.9").alanlar
        self.assertEqual(a["sipPbx"], "10.1.1.1")
        self.assertEqual(a["sipPbxKaynak"], "cihaz günlüğü")

    def test_kayitli_degilse_pbx_yazilmaz(self):
        """Kayıt olmamış cihaza PBX yazmak olmayan bağlantıyı var gösterir."""
        cihaz = self.kur_lcd()
        self.yamala(SahteAdb(logcat=(
            "08-05 12:04:17.249 I AnnounceSip: "
            "Registration state=failed code=408\n")))
        a = okuma.cihaz_oku(cihaz, pbx_ip="10.1.1.1").alanlar
        self.assertEqual(a["sipPbx"], "")
        self.assertEqual(a["sipPbxKaynak"], "")
        self.assertEqual(a["sipKayit"], "failed (408)")

    def test_broker_duyurusu_yoksa_numara_uydurulmaz(self):
        from core import piscu
        cihaz = self.kur_lcd()
        t = piscu.Telemetri("10.1.1.1")
        t.sip = {"10.9.9.9": "6001"}          # başka cihazın numarası
        self.yamala(SahteAdb(logcat=""))
        sonuc = okuma.cihaz_oku(cihaz, telemetri=t)
        self.assertEqual(sonuc.alanlar["sipDahili"], "")
        self.assertEqual(sonuc.alanlar["sipDahiliKaynak"], "")


class IpPlani(PanelTesti):

    def test_ip_atama_yalniz_intercom_grubunu_kabul_eder(self):
        desteklenen = [g["ad"] for g in kategori.GRUPLAR
                       if kategori.grup_islemi_destekler(g, "ip")]
        self.assertEqual(desteklenen, ["Intercom"])
        self.assertEqual(
            [g["ad"] for g in ip_atama.gruplari_coz(
                ["Intercom", "Handset", "Kamera", "Tümü"])],
            ["Intercom"])

    def test_port_ayristirma(self):
        izin = set(range(1, 25))
        self.assertEqual(ip_atama.portlar_ayristir("11-14", izin),
                         [11, 12, 13, 14])
        self.assertEqual(ip_atama.portlar_ayristir("11-13, 21 22", izin),
                         [11, 12, 13, 21, 22])
        self.assertEqual(ip_atama.metin_yap([11, 12, 13, 21]), "11-13, 21")
        for kotu in ("14-11", "abc", "1-", "-", "11;;;x"):
            with self.assertRaises(ValueError, msg=kotu):
                ip_atama.portlar_ayristir(kotu, izin)

    def test_izinsiz_port_reddedilir(self):
        with self.assertRaises(ValueError) as t:
            ip_atama.portlar_ayristir("11-14", {11, 12})
        self.assertIn("tanımlı olmayan port", str(t.exception))

    def test_fabrika_ip_sete_gore_degismez(self):
        """Cihaz kutudan çıkarken hangi sete gideceğini bilmiyor.

        Şablon çözülseydi set 8'de cihaz 10.8.1.12'de aranır ve hiç
        bulunamazdı.
        """
        harita = sahte.device_map([
            {"Name": "Intercom_1", "IP": "10.n.1.10", "IsActive": True,
             "Type": "Announcement", "SubType": "Intercom", "Port": "11",
             "Status": {}}], switch_ip="10.n.1.101")
        self.kur_harita(harita)
        for n in (1, 8, 112):
            env = device_map.yukle(n, self.harita_yolu)
            self.assertEqual(ip_atama.fabrika_ip(env), "10.1.1.12")
            plan = ip_atama.plan(env, "Intercom", [11])
            self.assertEqual(plan["satirlar"][0]["fabrika"], "10.1.1.12")
            # Hedef IP yine sete göre çözülür; değişmeyen yalnız fabrika.
            self.assertEqual(plan["satirlar"][0]["hedefIp"], f"10.{n}.1.10")

    def test_arama_adaylari_ag_ve_maskeden_cikar(self):
        adaylar = ip_atama.arama_adaylari("10.1.1.0", "255.255.255.0")
        self.assertEqual(adaylar[0], "10.1.1.1")
        self.assertEqual(adaylar[-1], "10.1.1.254")
        with self.assertRaises(ValueError) as t:
            ip_atama.arama_adaylari("10.0.0.0", "255.0.0.0")
        self.assertIn("çok geniş", str(t.exception))

    def test_arama_araligi_ag_maskenin_yerine_gecer(self):
        """Maske geniş olduğunda aranacak yer aralıkla daraltılır."""
        adaylar = ip_atama.arama_adaylari("10.0.0.0", "255.0.0.0",
                                          bas="10.1.1.10", son="10.1.1.12")
        self.assertEqual(adaylar, ["10.1.1.10", "10.1.1.11", "10.1.1.12"])

    def test_bozuk_arama_araligi_reddedilir(self):
        for bas, son, beklenen in (
                ("10.1.1.10", "", "birlikte"),
                ("10.1.1.60", "10.1.1.10", "küçük olamaz"),
                ("10.1.0.0", "10.1.9.0", "çok geniş")):
            with self.assertRaises(ValueError, msg=f"{bas}-{son}") as t:
                ip_atama.arama_adaylari("", "", bas=bas, son=son)
            self.assertIn(beklenen, str(t.exception))

    def test_kosu_iki_fabrika_adresini_de_dener(self):
        """Sabit fabrika adresi + sete göre çözülmüş adres, ikisi de aday.

        Fabrika adresi sabitlenmeden önce koşu cihazı 10.n.1.12'de
        arıyordu; sahadaki cihazların bir kısmı hâlâ orada. Yalnız sabit
        adrese bakmak onları "bulunamadı" yapıyordu.
        """
        harita = sahte.device_map([
            {"Name": "Intercom_1", "IP": "10.n.1.10", "IsActive": True,
             "Type": "Announcement", "SubType": "Intercom", "Port": "11",
             "Status": {}}], switch_ip="10.n.1.101")
        self.kur_harita(harita)
        env = device_map.yukle(8, self.harita_yolu)
        sw = env.switchler()[0]

        yakalanan = {}

        def sahte_calistir(mod, argv, satir_geri, iptal=None):
            yakalanan["argv"] = argv
            return 0

        eski = ip_atama._betigi_calistir
        ip_atama._betigi_calistir = sahte_calistir
        self.addCleanup(lambda: setattr(ip_atama, "_betigi_calistir", eski))

        satirlar = []
        ip_atama._intercom_kosu(env, sw, [11], ("admin", "x"),
                                satirlar.append, {})
        argv = yakalanan["argv"]
        self.assertEqual(argv[argv.index("--factory-ip") + 1], "10.1.1.12")
        self.assertIn("--default-ip", argv)
        self.assertIn("10.8.1.12", argv[argv.index("--default-ip") + 1:])
        self.assertTrue(any("10.1.1.12" in s and "10.8.1.12" in s
                            for s in satirlar), satirlar)
        # Maske cihaza dayatılmaz: cihazın kendi netmask'i korunur.
        self.assertNotIn("--force-netmask", argv)
        self.assertNotIn("--netmask", argv)

    def test_plan_devicemapten_cikar(self):
        harita = sahte.device_map([
            {"Name": "Intercom_1", "IP": "10.n.1.10", "IsActive": True,
             "Type": "Announcement", "SubType": "Intercom", "Port": "11",
             "Status": {}},
            {"Name": "Intercom_2", "IP": "10.n.1.11", "IsActive": True,
             "Type": "Announcement", "SubType": "Intercom", "Port": "12",
             "Status": {}},
        ], switch_ip="10.n.1.101")
        env = self.kur_harita(harita)
        plan = ip_atama.plan(env, "Intercom", [11, 12, 13])

        self.assertEqual(plan["hedefSayi"], 2)
        p11 = next(s for s in plan["satirlar"] if s["port"] == 11)
        self.assertEqual(p11["ad"], "Intercom_1")
        self.assertEqual(p11["hedefIp"], "10.1.1.10")
        p13 = next(s for s in plan["satirlar"] if s["port"] == 13)
        self.assertFalse(p13["uygulanabilir"])
        self.assertEqual(p13["ad"], "—")

    def test_korumali_port_hedefteyse_kosu_reddedilir(self):
        """Bilgisayarın portu ya da switch bağlantısı koşuya giremez.

        Koşu PoE'yi sırayla kapatıp açıyor; bu portlardan birine dokunursa
        kendi yolunu keser ve yarıda kalır.
        """
        harita = sahte.device_map([
            {"Name": "Intercom_1", "IP": "10.n.1.10", "IsActive": True,
             "Type": "Announcement", "SubType": "Intercom", "Port": "11",
             "Status": {}}], switch_ip="10.n.1.101")
        env = self.kur_harita(harita)
        for sebep in ("bilgisayar bu porta bağlı",
                      "diğer switch'e giden bağlantı"):
            with self.assertRaises(ValueError) as t:
                ip_atama.kosu(env, env.switchler()[0].id, [11],
                              lambda s: None, korumali={11: sebep})
            self.assertIn("kendi bağlantısını", str(t.exception))
            self.assertIn(sebep, str(t.exception))


class ExcelCiktisi(ServisTesti):

    def test_excel_yalniz_okunan_degerleri_yazar(self):
        """Şablon kopyalanır; okunmayan hücre boş kalır, şablon bozulmaz."""
        from core import ayar
        env = device_map.yukle(1, ayar.KOK / "DeviceMap.json")
        gor = isler.gorunum(1)

        anons = [c for c in env.cihazlar if c.yontem == "http"][0]
        s = dogrulama.basarili({"surum": "1.2.5", "seri": "SN-1",
                                "calisma": "01:00:00"}, "http")
        s.nesil = isler.sonraki_nesil()
        gor.yaz(anons.id, s)

        cikti = Path(tempfile.mkdtemp(prefix="excel-test-")) / "cikti.xlsx"
        sablon_once = ayar.EXCEL_SABLON.read_bytes()
        yol = excel.uret(env, gor.hepsi(), cikti=cikti)

        self.assertTrue(yol.is_file())
        self.assertGreater(yol.stat().st_size, 5000)
        # Şablon dosyasına dokunulmadı
        self.assertEqual(ayar.EXCEL_SABLON.read_bytes(), sablon_once)

        import openpyxl
        wb = openpyxl.load_workbook(yol)
        ws = wb["Kontrol Listesi"]
        dv = __import__("core.betik", fromlist=["betik"]).device_verify()
        sutun = {ws.cell(dv.HEADER_ROW, c).value: c
                 for c in range(1, ws.max_column + 1)
                 if ws.cell(dv.HEADER_ROW, c).value}

        okunan_satir = None
        for satir in range(dv.HEADER_ROW + 1, ws.max_row + 1):
            if ws.cell(satir, sutun[dv.COL_IP_TEMPLATE]).value == anons.ip_sablonu:
                okunan_satir = satir
                break
        self.assertIsNotNone(okunan_satir)
        self.assertEqual(ws.cell(okunan_satir, sutun["Versiyon"]).value, "1.2.5")
        self.assertEqual(ws.cell(okunan_satir, sutun["Bağlantı Bilgisi"]).value,
                         anons.ip)

        # Okunmayan bir cihazın sürüm hücresi BOŞ kalmalı
        digerleri = [c for c in env.cihazlar
                     if c.id != anons.id and gor.al(c.id) is None]
        self.assertTrue(digerleri)
        for satir in range(dv.HEADER_ROW + 1, ws.max_row + 1):
            tmpl = ws.cell(satir, sutun[dv.COL_IP_TEMPLATE]).value
            if tmpl in {c.ip_sablonu for c in digerleri}:
                self.assertIsNone(ws.cell(satir, sutun["Versiyon"]).value,
                                  f"{tmpl} için sahte sürüm yazılmış")


if __name__ == "__main__":
    unittest.main()
