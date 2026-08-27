#!/usr/bin/env python3
"""Field device verification.

Copies the Excel template, fills the SOFTWARE CHECK columns with data
collected from the devices in the field, and saves the result as a new file.
The template itself is never written to.

Data sources (docs/CIHAZ_ENDPOINTLERI.md):
  MQTT  ALFA/DeviceMap     -> fields shared by every device
  MQTT  ALFA/AppStatus/#   -> PISCU and HMI version + hardware id
  HTTP  /api/v1/system/... -> Announcement (Amplifier / Handset / Intercom)
  ISAPI /System/...        -> Camera, NVR
  HTTP  /stat/basicInfo    -> KYLAND switch
  ADB   getprop            -> LCD / Compartment

Columns are addressed by a stable ID, never by the heading printed in the
sheet (see COLUMN_HEADINGS). A reworded heading must not empty a column.

Usage:
    python3 device_verify.py                       # with the .env values
    python3 device_verify.py -n 3                  # set number 3
    python3 device_verify.py --only Camera NVR     # these types only
    python3 device_verify.py --dry-run             # stay off the network
    python3 device_verify.py --list                # print the target IPs
"""
from __future__ import annotations

import argparse
import concurrent.futures as cf
import json
import os
import re
import subprocess
import sys
import time
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
import requests
from requests.auth import HTTPBasicAuth, HTTPDigestAuth
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HERE = Path(__file__).resolve().parent

# ── column contract ──────────────────────────────────────────────────────
# id -> the heading that id carries in the shipped template (read from row 4).
# The panel keeps the same table in panel/checklist/columns.py; the two must
# be changed together.
COL_SECTION = "section"
COL_SWITCH = "switch"
COL_PORT = "port"
COL_DEVICE_DEFINITION = "deviceDefinition"
COL_IP_TEMPLATE = "ipTemplate"
COL_EXPECTED_IP = "expectedIp"
COL_EXPECTED_VERSION = "expectedVersion"
COL_EXPECTED_SIP_EXTENSION = "expectedSipExtension"
COL_DEVICE_NAME = "deviceName"
COL_CONNECTION_INFO = "connectionInfo"
COL_VERSION = "version"
COL_DEVICE_NUMBER = "deviceNumber"
COL_STATUS_DESCRIPTION = "statusDescription"
COL_UPTIME = "uptime"
COL_SPEAKER_VOLUME = "speakerVolume"
COL_MIC_VOLUME = "micVolume"
COL_SPEAKER_GAIN = "speakerGain"
COL_MIC_GAIN = "micGain"
COL_SIP_PBX = "sipPbx"
COL_SIP_EXTENSION = "sipExtension"
COL_SIP_OUTBOUND = "sipOutbound"
COL_TIMEZONE = "timezone"
COL_NETWORK_TIME = "networkTime"

COLUMN_HEADINGS = {
    COL_SECTION: "Section",
    COL_SWITCH: "Switch",
    COL_PORT: "Port",
    COL_DEVICE_DEFINITION: "Device definition",
    COL_IP_TEMPLATE: "IP template",
    COL_EXPECTED_IP: "Expected IP",
    COL_EXPECTED_VERSION: "Expected version",
    COL_EXPECTED_SIP_EXTENSION: "Expected SIP extension",
    COL_DEVICE_NAME: "Device name",
    COL_CONNECTION_INFO: "Connection info",
    COL_VERSION: "Version",
    COL_DEVICE_NUMBER: "Device number",
    COL_STATUS_DESCRIPTION: "Status description",
    COL_UPTIME: "Uptime",
    COL_SPEAKER_VOLUME: "Speaker volume",
    COL_MIC_VOLUME: "Microphone volume",
    COL_SPEAKER_GAIN: "Speaker gain",
    COL_MIC_GAIN: "Microphone gain",
    COL_SIP_PBX: "SIP PBX IP",
    COL_SIP_EXTENSION: "SIP extension",
    COL_SIP_OUTBOUND: "SIP outbound number",
    COL_TIMEZONE: "Time zone",
    COL_NETWORK_TIME: "Network/time check",
}
COLUMN_FOR_HEADING = {heading: column
                      for column, heading in COLUMN_HEADINGS.items()}

# Columns this script fills in; everything else is filled in by hand.
FILLABLE = [
    COL_DEVICE_NAME, COL_CONNECTION_INFO, COL_VERSION, COL_DEVICE_NUMBER,
    COL_STATUS_DESCRIPTION, COL_UPTIME, COL_SPEAKER_VOLUME, COL_MIC_VOLUME,
    COL_SPEAKER_GAIN, COL_MIC_GAIN, COL_SIP_PBX, COL_SIP_EXTENSION,
    COL_SIP_OUTBOUND, COL_TIMEZONE, COL_NETWORK_TIME,
]
HEADER_ROW = 4
NA_FILL = "FFE7E6E6"

# Values written into the status column; the summary formulas at the bottom
# of the sheet count them.
STATUS_ACTIVE = "Active"
STATUS_INACTIVE = "Inactive"
NETWORK_TIME_OK = "OK"

# Field names in device responses vary by manufacturer; pick() tries these
# candidates in turn (exact name -> trailing-segment match).
K_VERSION = ("firmwareversion", "firmware", "swversion", "softwareversion",
             "appversion", "fwversion", "version", "buildversion", "build")
K_SERIAL = ("serialnumber", "serialno", "serial", "sn", "devicesn",
            "deviceserial", "deviceid", "chipid", "uuid")
K_UPTIME = ("uptimeseconds", "uptimesec", "uptime", "systemuptime",
            "runtime", "runningtime")
K_SPEAKER = ("speakervolume", "speaker_volume", "speakerlevel", "spkvolume",
             "callvolume", "call_volume_out1", "outputvolume", "playvolume",
             "speaker", "volume")
K_MIC = ("microphonevolume", "microphone_volume", "micvolume", "mic_volume",
         "miclevel", "inputvolume", "recordvolume", "microphone", "mic")
K_SPEAKER_GAIN = ("speakergain", "speaker_gain", "spkgain", "outputgain",
                  "playgain", "amplifiergain")
K_MIC_GAIN = ("microphonegain", "microphone_gain", "micgain", "mic_gain",
              "inputgain", "recordgain")
K_OUTBOUND = ("pbxoutextension", "pbx_out_extension", "pbxoutext",
              "outextension", "out_extension", "outext",
              "outboundextension", "outbound_extension", "outboundext",
              "outbound_ext", "outboundcallextension", "outboundnumber",
              "outbound_number", "outboundcallnumber", "outboundcall",
              "call_on_input_0", "calloninput0", "destinationnumber",
              "destination", "dialnumber", "targetnumber", "callnumber",
              "callext", "outbound", "pbxout")

# API endpoints tried on Announcement devices (every JSON answer is merged)
ANNOUNCEMENT_ENDPOINTS = [
    "system/settings", "system/modes", "system/info", "system/status",
    "system/sip", "system/network", "system/audio",
    "sip", "sip/settings", "sip/config", "sip/status",
    "settings", "config", "network", "audio", "status", "info",
]
K_PBX = ("pbxip", "pbx_ip", "sippbx", "sip_pbx", "pbxserver", "sipserver",
         "pbxaddress", "pbxhost", "serverip", "sipserverip", "sipproxy",
         "registrar", "proxy", "pbx", "server", "host")
K_EXTENSION = ("pbxextension", "pbx_extension", "sipextension",
               "sip_extension", "extension", "ext")


# ---------------------------------------------------------------- helpers --
def load_env(path: Path) -> dict:
    """A simple .env reader (no external dependency)."""
    env = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def resolve(ip_template: str, set_no) -> str:
    """Replaces the 'n' in an IP template with the set number.

    10.n.1.24, set 3  ->  10.3.1.24
    """
    return re.sub(r"(?<![0-9a-zA-Z])n(?![0-9a-zA-Z])", str(set_no),
                  ip_template or "")


def flatten(obj, prefix="") -> dict:
    """Flattens nested JSON into a {lowercase.key: value} dictionary."""
    out = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}{k}".lower()
            if isinstance(v, (dict, list)):
                out.update(flatten(v, f"{key}."))
            else:
                out[key] = v
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            out.update(flatten(v, f"{prefix}{i}."))
    return out


def _norm(s: str) -> str:
    """sip_outbound-Extension -> sipoutboundextension"""
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def pick(flat: dict, *candidates, exclude=()):
    """Returns the first matching value among the candidate key names.

    Three passes, loosening as they go:
      1. exact key match             pbxIp            -> pbxip
      2. trailing-segment match      sip.pbxIp        -> pbxip
      3. substring match (>=7 chars) data.sipOutboundExtension -> outboundextension

    Keys carrying any fragment from `exclude` are dropped in pass 3, so a
    search for "extension" does not pick up "outboundExtension" by mistake.
    """
    def ok(v):
        return v not in ("", None) and str(v).strip() != ""

    for cand in candidates:                                   # 1
        v = flat.get(cand.lower())
        if ok(v):
            return v

    tails = {k: _norm(k.rsplit(".", 1)[-1]) for k in flat}
    for cand in candidates:                                   # 2
        nc = _norm(cand)
        for k, v in flat.items():
            if tails[k] == nc and ok(v):
                return v

    bad = tuple(_norm(x) for x in exclude)
    full = {k: _norm(k) for k in flat}
    for cand in candidates:                                   # 3
        nc = _norm(cand)
        if len(nc) < 7:
            continue
        for k, v in flat.items():
            nk = full[k]
            if nc in nk and ok(v) and not any(b in nk for b in bad):
                return v
    return None


def pct(v):
    if v in (None, ""):
        return None
    try:
        return f"{int(float(v))}%"
    except (TypeError, ValueError):
        return str(v)


def num(v):
    """Numeric fields such as gain — written to Excel as numbers (0 too)."""
    if v in (None, ""):
        return None
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return str(v)


def status_text(status: dict | None) -> str:
    """Active / Inactive. Fault flags are not broken out.

    The 'Has Network Failure' and 'Has Power Failure' flags in DeviceMap are
    not filled consistently (set on some powered-down devices and not on
    others), so no parenthesised reason is written.
    """
    if not status:
        return ""
    return STATUS_ACTIVE if status.get("NoError") else STATUS_INACTIVE


def uptime_text(seconds) -> str:
    """Converts seconds to HH:MM:SS (e.g. 03:25:41)."""
    try:
        u = int(float(seconds))
    except (TypeError, ValueError):
        return ""
    if u < 0:
        return ""
    return f"{u // 3600:02d}:{(u % 3600) // 60:02d}:{u % 60:02d}"


def xml_tag(root, tag):
    for el in root.iter():
        if el.tag.rsplit("}", 1)[-1] == tag:
            return (el.text or "").strip()
    return None


# Raw field names coming from the devices (written out with --debug-fields)
RAW: dict = {}

# Retained application status messages under ALFA/AppStatus/...
APP_STATUS: dict = {}          # ClientId -> payload


# ------------------------------------------------------------- collectors --
def fetch_device_map(broker: str, port: int, topic: str, timeout: float):
    """Reads the retained ALFA/DeviceMap message from PISCU."""
    try:
        import paho.mqtt.client as mqtt
        from paho.mqtt.enums import CallbackAPIVersion
    except ImportError:
        print("  [!] paho-mqtt is not installed, skipping the live DeviceMap "
              "(pip install paho-mqtt)")
        return None

    box, deadline = {}, timeout
    cl = mqtt.Client(CallbackAPIVersion.VERSION2,
                     client_id="commissioning_panel_verify")
    cl.on_connect = lambda c, u, f, rc, p=None: c.subscribe(topic)
    cl.on_message = lambda c, u, msg: box.setdefault("payload", msg.payload)
    try:
        cl.connect(broker, port, keepalive=10)
        cl.loop_start()
        for _ in range(int(deadline * 10)):
            if "payload" in box:
                break
            time.sleep(0.1)
        cl.loop_stop()
        cl.disconnect()
    except Exception as exc:
        print(f"  [!] MQTT: {exc}")
        return None
    if "payload" not in box:
        print(f"  [!] MQTT: no retained '{topic}' message arrived")
        return None
    return json.loads(box["payload"])


def fetch_app_status(broker: str, port: int, prefix: str, timeout: float) -> dict:
    """Collects the retained application messages under ALFA/AppStatus/#.

    Example payload: {"ClientId": "ClientManager_PISCU_YATAKLI_1",
                      "DeviceIP": ..., "HWID": ..., "Status": ...,
                      "Version": "1.2.7"}
    """
    try:
        import paho.mqtt.client as mqtt
        from paho.mqtt.enums import CallbackAPIVersion
    except ImportError:
        print("  [!] paho-mqtt is not installed, AppStatus cannot be read")
        return {}


    found = {}

    def on_message(client, userdata, msg):
        try:
            data = json.loads(msg.payload)
        except ValueError:
            return
        key = data.get("ClientId") or msg.topic.rsplit("/", 1)[-1]
        found[key] = data

    cl = mqtt.Client(CallbackAPIVersion.VERSION2,
                     client_id="commissioning_panel_appstatus")
    cl.on_connect = lambda c, u, f, rc, p=None: c.subscribe(f"{prefix}/#")
    cl.on_message = on_message
    try:
        cl.connect(broker, port, keepalive=10)
        cl.loop_start()
        time.sleep(timeout)               # retained messages arrive at once
        cl.loop_stop()
        cl.disconnect()
    except Exception as exc:
        print(f"  [!] AppStatus MQTT: {exc}")
        return {}
    return found


def app_status_for(ip: str, keyword: str) -> dict | None:
    """Finds the device's AppStatus record.

    DeviceIP is matched first. Falling back to the keyword (ClientId) is only
    valid when the message carries no DeviceIP at all — otherwise another
    set's record (10.1.1.4, say) would be written onto this set's row.
    """
    for data in APP_STATUS.values():
        if str(data.get("DeviceIP", "")).strip() == ip:
            return data
    kw = keyword.lower()
    for key, data in APP_STATUS.items():
        if kw in str(key).lower() and not str(data.get("DeviceIP", "")).strip():
            return data
    return None


def read_http_api(ip: str, cfg, tag: str) -> dict:
    """Walks the device's /api/v1/... endpoints and merges every JSON answer.

    Both Announcement devices and PISCU serve this API (version 1.2.x).
    """
    base = f"http://{ip}:{cfg.arduino_port}/api/v1"
    endpoints = cfg.announcement_endpoints or ANNOUNCEMENT_ENDPOINTS
    flat, seen, first_error = {}, {}, None
    for endpoint in endpoints:
        try:
            r = requests.get(f"{base}/{endpoint}", timeout=cfg.timeout)
            if not r.ok:
                continue
            part = flatten(r.json())
            seen[endpoint] = part
            flat.update(part)
        except Exception as exc:
            if first_error is None:
                first_error = exc
    if not flat:                     # no endpoint answered
        raise first_error or RuntimeError(f"the {tag} API did not answer")
    RAW.setdefault(ip, {})[tag] = seen
    return flat


def common_http_fields(flat: dict) -> dict:
    """The three fields every device shares: version, serial, uptime."""
    return {
        COL_VERSION:       pick(flat, *K_VERSION),
        COL_DEVICE_NUMBER: pick(flat, *K_SERIAL),
        COL_UPTIME:        uptime_text(pick(flat, *K_UPTIME)),
    }


def fetch_announcement(ip: str, cfg) -> dict:
    """Amplifier / Handset / Intercom — the Arduino-based HTTP API."""
    flat = read_http_api(ip, cfg, "announcement")
    return {
        COL_VERSION:       pick(flat, *K_VERSION),
        COL_DEVICE_NUMBER: pick(flat, *K_SERIAL),
        COL_UPTIME:        uptime_text(pick(flat, *K_UPTIME)),
        # "gain" and "level" fields are excluded so they cannot be mistaken
        # for the volume
        COL_SPEAKER_VOLUME: pct(pick(flat, *K_SPEAKER, exclude=("gain",))),
        COL_MIC_VOLUME:     pct(pick(flat, *K_MIC, exclude=("gain",))),
        COL_SPEAKER_GAIN:   num(pick(flat, *K_SPEAKER_GAIN)),
        COL_MIC_GAIN:       num(pick(flat, *K_MIC_GAIN)),
        COL_SIP_PBX:        pick(flat, *K_PBX),
        # excluded so a search for "extension" does not drift to
        # outboundExtension
        COL_SIP_EXTENSION:  pick(flat, *K_EXTENSION,
                                 exclude=("outbound", "outext",
                                          "pbxout", "dial", "target")),
        COL_SIP_OUTBOUND:   pick(flat, *K_OUTBOUND),
    }


def fetch_isapi(ip: str, cfg) -> dict:
    """Camera / NVR — Hikvision ISAPI (digest auth, XML)."""
    auth = HTTPDigestAuth(cfg.video_user, cfg.video_pass)
    base = f"http://{ip}:{cfg.video_port}/ISAPI"
    out, problems = {}, []

    r = requests.get(f"{base}/System/deviceInfo", auth=auth,
                     timeout=cfg.timeout, verify=False)
    root = ET.fromstring(r.content)
    out[COL_VERSION] = xml_tag(root, "firmwareVersion")
    out[COL_DEVICE_NUMBER] = xml_tag(root, "serialNumber")

    try:
        r = requests.get(f"{base}/System/time", auth=auth,
                         timeout=cfg.timeout, verify=False)
        if xml_tag(ET.fromstring(r.content), "timeZone") != cfg.expected_tz:
            problems.append("Time")
    except Exception:
        problems.append("Time")

    try:
        r = requests.get(f"{base}/System/time/ntpServers/1", auth=auth,
                         timeout=cfg.timeout, verify=False)
        if xml_tag(ET.fromstring(r.content), "ipAddress") != cfg.ntp_ip:
            problems.append("NTP")
    except Exception:
        problems.append("NTP")

    try:
        r = requests.get(f"{base}/System/Network/interfaces", auth=auth,
                         timeout=cfg.timeout, verify=False)
        root = ET.fromstring(r.content)
        mask = None
        for el in root.iter():
            if el.tag.rsplit("}", 1)[-1] == "IPAddress":
                addr = sub = None
                for ch in el:
                    tag = ch.tag.rsplit("}", 1)[-1]
                    if tag == "ipAddress":
                        addr = (ch.text or "").strip()
                    elif tag == "subnetMask":
                        sub = (ch.text or "").strip()
                if addr and addr.startswith("10."):
                    mask = sub
                    break
        if mask != cfg.expected_mask:
            problems.append("Mask")
    except Exception:
        problems.append("Mask")

    out[COL_NETWORK_TIME] = (NETWORK_TIME_OK if not problems
                             else ", ".join(problems))
    return out


def fetch_switch(ip: str, cfg) -> dict:
    """KYLAND switch — HTTP basic auth."""
    r = requests.get(f"http://{ip}:{cfg.kyland_port}/stat/basicInfo",
                     auth=HTTPBasicAuth(cfg.kyland_user, cfg.kyland_pass),
                     timeout=cfg.timeout)
    r.raise_for_status()
    try:
        flat = flatten(r.json())
    except ValueError:                       # not JSON: parse the raw text
        flat = flatten(dict(re.findall(r'"?([A-Za-z_]+)"?\s*[:=]\s*"?([^",\n}]+)',
                                       r.text)))
    RAW.setdefault(ip, {})["switch"] = flat
    out = common_http_fields(flat)
    out.pop(COL_DEVICE_NUMBER, None)     # the switch serial is not collected
    return out


def app_status_fields(ip: str, cfg, keyword: str) -> dict:
    """Firmware version and hardware id from the ALFA/AppStatus message.

    Devices running the application (PISCU, HMI) publish their own state
    here:
      ClientManager_PISCU_YATAKLI_1  ip=10.n.1.1  v=1.2.7  hwid=604A17F3
      ClientManager_MCP_YATAKLI_1    ip=10.n.1.4  v=1.2.5  hwid=34DA8534
    """
    rec = app_status_for(ip, keyword)
    if not rec:
        raise RuntimeError(f"no {keyword} message found under "
                           f"{cfg.app_status_prefix}/...")
    RAW.setdefault(ip, {})[f"appstatus_{keyword.lower()}"] = rec
    return {
        COL_VERSION:       rec.get("Version"),
        COL_DEVICE_NUMBER: rec.get("HWID"),
    }


def fetch_piscu(ip: str, cfg) -> dict:
    """PISCU — the ClientManager_PISCU_* application status."""
    return app_status_fields(ip, cfg, "PISCU")


def fetch_hmi(ip: str, cfg) -> dict:
    """HMI — the ClientManager_MCP_* application status (no SSH needed)."""
    return app_status_fields(ip, cfg, "MCP")


def _adb(target: str, *args, timeout: int) -> str:
    r = subprocess.run(["adb", "-s", target, *args],
                       capture_output=True, text=True, timeout=timeout)
    return r.stdout.strip()


def fetch_compartment_lcd(ip: str, cfg) -> dict:
    """LCD / Compartment — Android, over ADB."""
    target = f"{ip}:{cfg.adb_port}"
    subprocess.run(["adb", "connect", target], capture_output=True,
                   text=True, timeout=cfg.adb_timeout)
    try:
        # NOTE: the version is not collected for now — grey in the template.
        # ro.build.display.id is the Android build id (C33P-V1.5-...), and
        # DeviceMap Status.Version is not the right source either. This will
        # be added once it is clear where the app version should be read
        # from.
        out = {
            COL_DEVICE_NUMBER: _adb(target, "shell", "getprop", "ro.serialno",
                                    timeout=cfg.adb_timeout),
            COL_TIMEZONE:      _adb(target, "shell", "getprop",
                                    "persist.sys.timezone",
                                    timeout=cfg.adb_timeout),
        }
        raw = _adb(target, "shell", "cat", "/proc/uptime", timeout=cfg.adb_timeout)
        if raw:
            out[COL_UPTIME] = uptime_text(raw.split()[0])
        return {k: v for k, v in out.items() if v}
    finally:
        subprocess.run(["adb", "disconnect", target], capture_output=True,
                       text=True, timeout=cfg.adb_timeout)


# Which type uses which collector (an extra query beyond DeviceMap)
COLLECTORS = {
    ("Switch", None):              fetch_switch,
    ("PISCU", None):               fetch_piscu,
    ("HMI", None):                 fetch_hmi,
    ("Announcement", "Amplifier"): fetch_announcement,
    ("Announcement", "Handset"):   fetch_announcement,
    ("Announcement", "Intercom"):  fetch_announcement,
    ("Camera", "Corridor"):        fetch_isapi,
    ("Camera", "Landing"):         fetch_isapi,
    ("NVR", None):                 fetch_isapi,
    ("LCD", "Compartment"):        fetch_compartment_lcd,
}
# Fed from DeviceMap only, with no extra query: ICU, AP, LED, LCD/Landing,
# and Announcement/UIC. (PISCU and HMI are NOT in that list — they are
# collected above, over MQTT AppStatus.)
#
# The UIC is the one the panel and this table see differently: the panel
# WRITES its SIP and threshold settings over HTTP, so it carries an `http`
# read method there, while everything the checklist reports for a UIC
# already arrives in DeviceMap. tests/test_data.py holds that difference
# with its reason; if a collector is added here, the exemption there must go.


# ------------------------------------------------------------------- flow --
def build_index(device_map: dict) -> dict:
    """IP template -> device record."""
    idx = {}
    for sw in device_map.get("Switches", []) or []:
        idx.setdefault("__trainset__", sw.get("TrainSet"))
        idx[sw.get("IP", "")] = {
            "Name": sw.get("Name", ""), "Type": "Switch", "SubType": None,
            "Status": sw.get("Status", {}), "SerialNumber": sw.get("SerialNumber", ""),
            "PBXExtension": None,
        }
        for dv in sw.get("Devices", []) or []:
            idx[dv.get("IP", "")] = {
                "Name": dv.get("Name", ""), "Type": dv.get("Type", ""),
                "SubType": dv.get("SubType") or None,
                "Status": dv.get("Status", {}),
                "SerialNumber": dv.get("SerialNumber", ""),
                "PBXExtension": dv.get("PBXExtension") or None,
            }
    return idx


def map_train_set(index: dict):
    """Which set this DeviceMap belongs to (None when absent)."""
    return index.get("__trainset__")


def devices_only(index: dict) -> dict:
    return {k: v for k, v in index.items() if k != "__trainset__"}


def remove_file(path: Path, tries: int = 4) -> bool:
    """Deletes a file persistently. On systems that will not delete an open
    file it renames first and deletes that; failing everything, it truncates
    the file to zero bytes."""

    for attempt in range(tries):
        if not path.exists():
            return True
        try:
            path.unlink()
            return True
        except OSError:
            pass
        try:                                  # move the locked file aside
            tmp = path.with_name(f"{path.name}.old{os.getpid()}{attempt}")
            path.rename(tmp)
            try:
                tmp.unlink()
            except OSError:
                pass
            return True
        except OSError:
            time.sleep(0.3)
    try:                                      # last resort: empty it
        with open(path, "wb"):
            pass
        return True
    except OSError:
        return False


def clear_output(cfg, out_path: Path) -> bool:
    """Removes the output and any office lock files next to it."""
    for lock in (out_path.with_name(f".~lock.{out_path.name}#"),
                 out_path.with_name(f"~${out_path.name}")):
        if lock.exists():
            remove_file(lock)
    if remove_file(out_path):
        return True
    print(f"\n[ERROR] Could not delete the old output: {out_path.name}")
    print("        The file may be open in Excel / LibreOffice.")
    print("        Close it and run the script again.")
    return False


def prepare_output(cfg, n) -> Path | None:
    """Works out the output path and deletes the old file up front.

    Called before going onto the network — if the file cannot be deleted, that
    is clear immediately instead of after a three-minute wait.
    """
    out_path = cfg.output or cfg.template.with_name(
        f"{cfg.template.stem}_set{n}{cfg.template.suffix}")
    if cfg.dry_run or cfg.list:
        return out_path
    return out_path if clear_output(cfg, out_path) else None


def parse_args(env: dict):
    p = argparse.ArgumentParser(
        description="Field device verification — Excel filler",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)

    g = p.add_argument_group("files")
    g.add_argument("--template", type=Path,
                   default=HERE / "Field_Device_Verification.xlsx",
                   help="template Excel (never written over)")
    g.add_argument("--output", type=Path, default=None,
                   help="output file (default: template_set<N>.xlsx, "
                        "overwritten if it exists)")
    g.add_argument("--device-map", type=Path,
                   default=Path(env.get("DEVICE_MAP_FILE") or HERE / "DeviceMap.json"),
                   help="local DeviceMap.json")
    g.add_argument("--sheet", default="Checklist")

    g = p.add_argument_group("set / network")
    g.add_argument("-n", "--set", dest="set_no",
                   default=env.get("TRAIN_SET_NO", "1"),
                   help="set number — the 'n' in the IP template (2nd octet)")
    g.add_argument("--piscu-ip", default=None,
                   help="MQTT broker IP (default: the PISCU in DeviceMap)")
    g.add_argument("--mqtt-port", type=int, default=int(env.get("PISCU_MQTT_PORT", 1883)))
    g.add_argument("--mqtt-topic", default=env.get("PISCU_DEVICE_MAP_TOPIC", "ALFA/DeviceMap"))
    g.add_argument("--app-status-prefix",
                   default=env.get("PISCU_APP_STATUS_PREFIX", "ALFA/AppStatus"),
                   help="application status topic prefix")
    g.add_argument("--app-status-timeout", type=float, default=3.0)
    g.add_argument("--no-mqtt", action="store_true",
                   help="do not fetch the live DeviceMap; use the local file")

    g = p.add_argument_group("ports / credentials")
    g.add_argument("--arduino-port", type=int, default=int(env.get("ARDUINO_HTTP_PORT", 80)))
    g.add_argument("--video-port", type=int, default=int(env.get("VIDEO_HTTP_PORT", 80)))
    g.add_argument("--kyland-port", type=int, default=int(env.get("KYLAND_HTTP_PORT", 80)))
    g.add_argument("--adb-port", type=int, default=int(env.get("COMPARTMENT_LCD_ADB_PORT", 5555)))
    g.add_argument("--video-user", default=env.get("VIDEO_USERNAME", "admin"))
    g.add_argument("--video-pass", default=env.get("VIDEO_PASSWORD", ""))
    g.add_argument("--kyland-user", default=env.get("KYLAND_USERNAME", "admin"))
    g.add_argument("--kyland-pass", default=env.get("KYLAND_PASSWORD", ""))

    g = p.add_argument_group("expected values")
    g.add_argument("--expected-tz", default=env.get("EXPECTED_TIMEZONE", "CST-3:00:00"),
                   help="camera/NVR time zone")
    g.add_argument("--ntp-ip", default=None,
                   help="expected NTP server (default: the PISCU IP)")
    # The sleeper-coach devices report 255.255.0.0 (it was 255.0.0.0 on the
    # earlier project).
    g.add_argument("--expected-mask",
                   default=env.get("EXPECTED_SUBNET_MASK", "255.255.0.0"),
                   help="expected subnet mask for camera/NVR")

    g = p.add_argument_group("run")
    g.add_argument("--timeout", type=float, default=5.0, help="HTTP timeout (s)")
    g.add_argument("--adb-timeout", type=int, default=15)
    g.add_argument("--mqtt-timeout", type=float, default=5.0)
    g.add_argument("--workers", type=int, default=12,
                   help="number of concurrent queries")
    g.add_argument("--only", nargs="+", metavar="TYPE",
                   help="query these Type values only (e.g. Camera NVR)")
    g.add_argument("--skip", nargs="+", metavar="TYPE", default=[],
                   help="skip these Type values")
    g.add_argument("--dry-run", action="store_true",
                   help="stay off the network; show the plan")
    g.add_argument("--list", action="store_true",
                   help="print the target IP list and exit")
    g.add_argument("--debug-fields", action="store_true",
                   help="save the raw field names from the devices as JSON")
    g.add_argument("--announcement-endpoints", nargs="+", metavar="PATH",
                   default=None,
                   help="/api/v1/<path> endpoints tried on Announcement "
                        f"devices (default: {' '.join(ANNOUNCEMENT_ENDPOINTS)})")

    return p.parse_args()


def main() -> int:
    env = load_env(HERE / ".env")
    env.update({k: v for k, v in os.environ.items() if k in env})
    cfg = parse_args(env)

    if not cfg.template.exists():
        print(f"[ERROR] Template not found: {cfg.template}")
        return 1
    if not cfg.device_map.exists():
        print(f"[ERROR] DeviceMap not found: {cfg.device_map}")
        return 1

    n = cfg.set_no
    local_map = json.loads(cfg.device_map.read_text(encoding="utf-8"))
    index = build_index(local_map)
    map_is_live = False              # does the telemetry belong to this set?

    # PISCU / NTP defaults
    piscu_tmpl = next((ip for ip, d in devices_only(index).items()
                       if d["Type"] == "PISCU"), None)
    cfg.piscu_ip = cfg.piscu_ip or (resolve(piscu_tmpl, n) if piscu_tmpl else None)
    cfg.ntp_ip = cfg.ntp_ip or cfg.piscu_ip

    print(f"Set no           : {n}")
    print(f"PISCU / broker   : {cfg.piscu_ip}")
    print(f"Template         : {cfg.template.name}")
    print(f"DeviceMap        : {cfg.device_map.name}  "
          f"({len(devices_only(index))} records, set "
          f"{map_train_set(index) or '?'})")

    # Prepare the output file first — if it is open, do not bother with the
    # network at all
    out_path = prepare_output(cfg, n)
    if out_path is None:
        return 1
    if not (cfg.dry_run or cfg.list):
        print(f"Output           : {out_path.name}")

    # 1) live DeviceMap
    if not cfg.no_mqtt and not cfg.dry_run and not cfg.list and cfg.piscu_ip:
        print("\n[1/3] Live DeviceMap (MQTT)...")
        live = fetch_device_map(cfg.piscu_ip, cfg.mqtt_port,
                                cfg.mqtt_topic, cfg.mqtt_timeout)
        if live:
            index = build_index(live)
            ts = map_train_set(index)
            print(f"  -> {len(devices_only(index))} records received "
                  f"(set {ts or '?'})")
            if ts is not None and str(ts) != str(n):
                print(f"  [!] The broker reports set {ts}, {n} was asked for. "
                      f"The telemetry will not be used.")
            else:
                map_is_live = True
        else:
            print("  -> no live telemetry")

        APP_STATUS.update(fetch_app_status(cfg.piscu_ip, cfg.mqtt_port,
                                           cfg.app_status_prefix,
                                           cfg.app_status_timeout))
        print(f"  AppStatus: {len(APP_STATUS)} application message(s) "
              f"({cfg.app_status_prefix}/#)")
        for key, data in sorted(APP_STATUS.items()):
            print(f"    · {key:<34} ip={data.get('DeviceIP', '—'):<12} "
                  f"v={data.get('Version', '—'):<8} "
                  f"hwid={data.get('HWID', '—'):<12} {data.get('Status', '')}")

    # 2) Load the template into memory — the file is touched only at the end
    wb = openpyxl.load_workbook(cfg.template)
    ws = wb[cfg.sheet]

    # Heading -> id -> index. Headings live in the sheet; ids live in the
    # code, and only ids are used below.
    col = {}
    for index_ in range(1, ws.max_column + 1):
        column = COLUMN_FOR_HEADING.get(
            str(ws.cell(HEADER_ROW, index_).value or "").strip())
        if column:
            col[column] = index_
    missing = [c for c in FILLABLE + [COL_IP_TEMPLATE] if c not in col]
    if missing:
        print(f"[ERROR] Column not found in the template: {missing}")
        return 1

    # 3) work out the targets
    targets = []
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        tmpl = ws.cell(row, col[COL_IP_TEMPLATE]).value
        if not tmpl or not str(tmpl).startswith("10."):
            continue
        dev = index.get(str(tmpl))
        if not dev:
            continue
        typ, sub = dev["Type"], dev["SubType"]
        if cfg.only and typ not in cfg.only:
            continue
        if typ in cfg.skip:
            continue
        targets.append((row, str(tmpl), resolve(str(tmpl), n), dev))

    if cfg.list or cfg.dry_run:
        print(f"\nTargets ({len(targets)}):")
        for row, tmpl, ip, dev in targets:
            fn = COLLECTORS.get((dev["Type"], dev["SubType"]))
            src = fn.__name__.replace("fetch_", "") if fn else "DeviceMap only"
            kind = f"{dev['Type']}/{dev['SubType'] or '-'}"
            print(f"  row {row:>3}  {ip:<14} {kind:<26} "
                  f"{dev['Name']:<22} <- {src}")
        print(f"\nThe output would have been: {out_path.name}")
        return 0

    # 4) shared fields (DeviceMap) + type-specific queries
    print(f"\n[2/3] Querying {len(targets)} device(s) "
          f"({cfg.workers} at a time)...")

    def work(target):
        row, tmpl, ip, dev = target
        # The device name is identity data and is always written.
        values = {COL_DEVICE_NAME: dev["Name"]}

        # Status fields are written ONLY if the telemetry belongs to this
        # set. Otherwise another set's data (10.1.1.x, say) leaks into the
        # 10.2.1.x rows.
        if map_is_live:
            status = dev.get("Status") or {}
            values[COL_STATUS_DESCRIPTION] = status_text(status)
            # PISCU remembers the LAST KNOWN version/serial/uptime of
            # powered-down devices. Those do not reflect the present, so they
            # are written only while the device is Active — otherwise an
            # unplugged intercom looks like it is reporting a version.
            if status.get("NoError"):
                values.update({
                    COL_CONNECTION_INFO: ip,
                    COL_DEVICE_NUMBER:   dev.get("SerialNumber") or "",
                    COL_UPTIME:          uptime_text(status.get("Uptime")),
                    COL_VERSION:         status.get("Version") or "",
                })
        # NOTE: PBXExtension in DeviceMap is a *definition* value, not one
        # read from the device. It belongs in the "Expected SIP extension"
        # column of the template; only what the device reports itself is
        # written here.
        fn = COLLECTORS.get((dev["Type"], dev["SubType"]))
        err = None
        if fn:
            try:
                got = fn(ip, cfg) or {}
                if any(v not in (None, "") for v in got.values()):
                    values[COL_CONNECTION_INFO] = ip   # the device did answer
                for k, v in got.items():
                    if v not in (None, ""):
                        values[k] = v
            except Exception as exc:
                err = f"{type(exc).__name__}: {exc}"
        return row, ip, dev, values, err

    results, errors = [], []
    with cf.ThreadPoolExecutor(max_workers=cfg.workers) as pool:
        for row, ip, dev, values, err in pool.map(work, targets):
            results.append((row, values))
            if err:
                errors.append((ip, dev["Name"], err))

    # 5) Write to Excel — only into columns valid for that device type
    written = 0
    for row, values in results:
        for column, value in values.items():
            if column not in FILLABLE or value in (None, ""):
                continue
            cell = ws.cell(row, col[column])
            if (cell.fill and cell.fill.fgColor
                    and cell.fill.fgColor.rgb == NA_FILL):
                continue                 # grey = invalid field for this type
            cell.value = value
            written += 1

    # 6) Delete the file once more right before writing, then create it fresh
    if not clear_output(cfg, out_path):
        return 1
    wb.save(out_path)

    if cfg.debug_fields:
        dbg = out_path.with_suffix(".fields.json")
        dbg.write_text(json.dumps(RAW, ensure_ascii=False, indent=2, default=str),
                       encoding="utf-8")
        print(f"  raw field names -> {dbg.name}")

    print(f"\n[3/3] Saved: {out_path.name}")
    print(f"  {len(results)} rows, {written} cells filled")
    if errors:
        print(f"  {len(errors)} device(s) unreachable:")
        for ip, name, err in errors:
            print(f"    - {ip:<14} {name:<22} {err[:70]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
